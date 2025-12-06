from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework_simplejwt.tokens import RefreshToken
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.db.models import Q, Count
from django.http import HttpResponse
from datetime import timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io
import logging

# Audit logging
from .utils.audit import AuditLogger, AuditAction

audit_logger = logging.getLogger('nexpro.audit')
from .models import (
    Organization, OrganizationEmail, Subscription,
    Client, WorkType, WorkTypeAssignment, ClientWorkMapping, WorkInstance,
    EmailTemplate, ReminderRule, ReminderInstance, Notification, TaskDocument,
    ReportConfiguration, EmailOTP, AuditLog, PlatformSettings, SubscriptionPlan,
    CredentialVault, GoogleConnection, GoogleSyncSettings, GoogleSyncLog,
    GoogleTaskMapping, GoogleCalendarMapping, GoogleDriveMapping, GoogleAPIQuotaUsage,
    SubTaskCategory
)
from .serializers import (
    OrganizationSerializer, OrganizationMinimalSerializer,
    OrganizationRegistrationSerializer, SubscriptionSerializer,
    OrganizationEmailSerializer, OrganizationEmailWriteSerializer,
    UserSerializer, ClientSerializer, WorkTypeSerializer,
    WorkTypeAssignmentSerializer, SubTaskCategorySerializer,
    ClientWorkMappingSerializer, WorkInstanceSerializer,
    EmailTemplateSerializer,
    ReminderRuleSerializer, ReminderInstanceSerializer,
    NotificationSerializer, CustomTokenObtainPairSerializer,
    TaskDocumentSerializer, ReportConfigurationSerializer,
    PlatformSettingsSerializer, SubscriptionPlanSerializer,
    CredentialVaultSerializer, CredentialVaultDecryptedSerializer,
    GoogleConnectionSerializer, GoogleConnectionUpdateSerializer,
    GoogleSyncSettingsSerializer, GoogleSyncLogSerializer,
    GoogleTaskMappingSerializer, GoogleCalendarMappingSerializer,
    GoogleDriveMappingSerializer, GoogleAPIQuotaUsageSerializer
)
from .permissions import (
    IsPlatformAdmin, IsOrganizationAdmin, IsAdminOrPartner,
    IsAdminOrReadOnly, IsSameOrganization, CanManageUsers,
    IsOrganizationActive
)
from .services.task_service import TaskAutomationService
from .services.email_service import EmailService
from .services.plan_service import PlanService
from .services.otp_service import OTPService

User = get_user_model()


# =============================================================================
# TENANT VIEWSET MIXIN
# =============================================================================

class TenantViewSetMixin:
    """
    Mixin to ensure tenant isolation in ViewSets.
    Automatically filters querysets by organization and sets organization on create.
    """

    def get_queryset(self):
        """Filter queryset by current user's organization"""
        qs = super().get_queryset()
        user = self.request.user

        # Platform admins can see all if requested
        if getattr(user, 'is_platform_admin', False):
            org_id = self.request.headers.get('X-Organization-ID')
            if org_id:
                return qs.filter(organization_id=org_id)
            return qs

        # Regular users: filter by their organization
        if hasattr(self.request, 'organization') and self.request.organization:
            return qs.filter(organization=self.request.organization)

        return qs.none()

    def perform_create(self, serializer):
        """Set organization when creating objects"""
        if hasattr(self.request, 'organization') and self.request.organization:
            serializer.save(organization=self.request.organization)
        else:
            serializer.save()


# =============================================================================
# AUTHENTICATION VIEWS
# =============================================================================

class CustomTokenObtainPairView(TokenObtainPairView):
    """Custom login view that returns user and organization data along with tokens"""
    serializer_class = CustomTokenObtainPairSerializer


class SendSignupOTPView(APIView):
    """
    Step 1 of registration: Send OTP for email verification.
    Only requires email - other details will be collected after OTP verification.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email')

        if not email:
            return Response({
                'error': 'Email is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validate email format
        import re
        email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not re.match(email_pattern, email):
            return Response({
                'error': 'Invalid email format'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Check if email already exists as a user
        if User.objects.filter(email=email).exists():
            return Response({
                'error': 'A user with this email already exists. Please login instead.'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Check if email already exists as organization email
        if Organization.objects.filter(email=email).exists():
            return Response({
                'error': 'An organization with this email already exists'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Store minimal signup data - just the email for now
        signup_data = {
            'admin_email': email,
            'email_verified': False,
        }

        # Send OTP
        success, message = OTPService.send_signup_otp(email, signup_data, request)

        if success:
            return Response({
                'message': message,
                'email': email
            }, status=status.HTTP_200_OK)

        return Response({'error': message}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class VerifySignupOTPView(APIView):
    """
    Step 2 of registration: Verify OTP only.
    Returns a verification token that must be used in step 3 to complete registration.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email')
        otp_code = request.data.get('otp')

        if not email or not otp_code:
            return Response({
                'error': 'Email and OTP are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Verify OTP
        success, result = OTPService.verify_otp(email, otp_code, 'SIGNUP', request)

        if not success:
            return Response({'error': result}, status=status.HTTP_400_BAD_REQUEST)

        # OTP verified - update signup data to mark email as verified
        otp_record = result
        signup_data = otp_record.signup_data or {}
        signup_data['email_verified'] = True
        signup_data['admin_email'] = email
        otp_record.signup_data = signup_data
        otp_record.is_verified = True
        otp_record.save()

        # Generate a verification token (using the OTP record ID as reference)
        import hashlib
        import secrets
        verification_token = hashlib.sha256(f"{otp_record.id}{secrets.token_hex(16)}".encode()).hexdigest()

        # Store verification token in signup_data
        signup_data['verification_token'] = verification_token
        otp_record.signup_data = signup_data
        otp_record.save()

        return Response({
            'message': 'Email verified successfully! Please complete your registration.',
            'email': email,
            'verified': True,
            'verification_token': verification_token
        }, status=status.HTTP_200_OK)


class CompleteSignupView(APIView):
    """
    Step 3 of registration: Complete registration with all details.
    Requires verified email and all organization/user details.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email')
        verification_token = request.data.get('verification_token')

        if not email or not verification_token:
            return Response({
                'error': 'Email and verification token are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Find the verified OTP record
        try:
            otp_record = EmailOTP.objects.get(
                email=email,
                purpose='SIGNUP',
                is_verified=True
            )
        except EmailOTP.DoesNotExist:
            return Response({
                'error': 'Email not verified. Please verify your email first.'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Verify the verification token
        signup_data = otp_record.signup_data or {}
        if signup_data.get('verification_token') != verification_token:
            return Response({
                'error': 'Invalid verification token. Please restart registration.'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Validate required fields for registration
        required_fields = [
            'organization_name', 'organization_phone', 'admin_first_name', 'admin_password'
        ]
        missing_fields = [f for f in required_fields if not request.data.get(f)]
        if missing_fields:
            return Response({
                'error': f'Missing required fields: {", ".join(missing_fields)}'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Use the verified email as both admin email and organization email
        org_email = request.data.get('organization_email', email)

        # Check if organization email already exists (if different from admin email)
        if org_email != email and Organization.objects.filter(email=org_email).exists():
            return Response({
                'error': 'An organization with this email already exists'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Create organization and user using the serializer
        serializer = OrganizationRegistrationSerializer(data={
            'organization_name': request.data.get('organization_name'),
            'organization_email': org_email,
            'organization_phone': request.data.get('organization_phone', ''),
            'admin_first_name': request.data.get('admin_first_name'),
            'admin_last_name': request.data.get('admin_last_name', ''),
            'admin_email': email,  # Use verified email
            'admin_password': request.data.get('admin_password'),
        })

        if serializer.is_valid():
            result = serializer.save()
            organization = result['organization']
            user = result['user']

            # Log successful registration
            AuditLog.log(
                action='SIGNUP',
                user=user,
                organization=organization,
                description=f'New organization registered: {organization.name}',
                request=request
            )

            # Send welcome email to user
            OTPService.send_welcome_email(organization, user)

            # Send notification to platform admins
            OTPService.send_admin_notification(organization, user)

            # Generate tokens for the new user
            refresh = RefreshToken.for_user(user)

            # Clean up OTP record
            otp_record.delete()

            return Response({
                'message': 'Registration successful! Welcome to NexPro.',
                'organization': OrganizationMinimalSerializer(organization).data,
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'role': user.role,
                },
                'tokens': {
                    'access': str(refresh.access_token),
                    'refresh': str(refresh),
                }
            }, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ResendSignupOTPView(APIView):
    """
    Resend OTP for signup email verification.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email')

        if not email:
            return Response({
                'error': 'Email is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Check if there's a pending OTP for this email
        try:
            existing_otp = EmailOTP.objects.get(
                email=email,
                purpose='SIGNUP',
                is_verified=False
            )
            signup_data = existing_otp.signup_data
        except EmailOTP.DoesNotExist:
            return Response({
                'error': 'No pending registration found for this email. Please start registration again.'
            }, status=status.HTTP_400_BAD_REQUEST)

        if not signup_data:
            return Response({
                'error': 'Registration data not found. Please start registration again.'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Resend OTP
        success, message = OTPService.send_signup_otp(email, signup_data, request)

        if success:
            return Response({
                'message': 'OTP resent successfully',
                'email': email
            }, status=status.HTTP_200_OK)

        return Response({'error': message}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RegisterView(APIView):
    """
    Legacy registration endpoint (kept for backwards compatibility).
    Redirects to OTP-based flow in production.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        serializer = OrganizationRegistrationSerializer(data=request.data)

        if serializer.is_valid():
            result = serializer.save()
            organization = result['organization']
            user = result['user']

            # Log registration
            AuditLog.log(
                action='SIGNUP',
                user=user,
                organization=organization,
                description=f'Organization registered (legacy flow): {organization.name}',
                request=request
            )

            # Send welcome email
            OTPService.send_welcome_email(organization, user)

            # Send admin notification
            OTPService.send_admin_notification(organization, user)

            # Generate tokens for the new user
            refresh = RefreshToken.for_user(user)

            return Response({
                'message': 'Registration successful',
                'organization': OrganizationMinimalSerializer(organization).data,
                'user': {
                    'id': user.id,
                    'email': user.email,
                    'first_name': user.first_name,
                    'last_name': user.last_name,
                    'role': user.role,
                },
                'tokens': {
                    'access': str(refresh.access_token),
                    'refresh': str(refresh),
                }
            }, status=status.HTTP_201_CREATED)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ForgotPasswordView(APIView):
    """
    Step 1 of password reset: Send OTP for verification.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email')

        if not email:
            return Response({
                'error': 'Email is required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Send OTP (will return generic message even if user doesn't exist)
        success, message = OTPService.send_password_reset_otp(email, request)

        return Response({
            'message': message
        }, status=status.HTTP_200_OK)


class VerifyPasswordResetOTPView(APIView):
    """
    Step 2 of password reset: Verify OTP.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        email = request.data.get('email')
        otp_code = request.data.get('otp')

        if not email or not otp_code:
            return Response({
                'error': 'Email and OTP are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        # Verify OTP
        success, result = OTPService.verify_otp(email, otp_code, 'PASSWORD_RESET', request)

        if not success:
            return Response({'error': result}, status=status.HTTP_400_BAD_REQUEST)

        # Return a reset token (the OTP ID) for the next step
        otp_record = result
        return Response({
            'message': 'OTP verified successfully',
            'reset_token': str(otp_record.id)
        }, status=status.HTTP_200_OK)


class ResetPasswordView(APIView):
    """
    Step 3 of password reset: Set new password.
    """
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        reset_token = request.data.get('reset_token')
        new_password = request.data.get('new_password')
        confirm_password = request.data.get('confirm_password')

        if not reset_token or not new_password or not confirm_password:
            return Response({
                'error': 'Reset token, new password and confirmation are required'
            }, status=status.HTTP_400_BAD_REQUEST)

        if new_password != confirm_password:
            return Response({
                'error': 'Passwords do not match'
            }, status=status.HTTP_400_BAD_REQUEST)

        if len(new_password) < 8:
            return Response({
                'error': 'Password must be at least 8 characters long'
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Find the verified OTP record
            otp_record = EmailOTP.objects.get(
                id=reset_token,
                purpose='PASSWORD_RESET',
                is_verified=True
            )

            # Check if reset was done within 15 minutes of OTP verification
            time_since_verification = timezone.now() - otp_record.verified_at
            if time_since_verification.total_seconds() > 900:  # 15 minutes
                return Response({
                    'error': 'Reset session expired. Please request a new OTP.'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get the user
            user = otp_record.user
            if not user:
                return Response({
                    'error': 'User not found. Please contact support.'
                }, status=status.HTTP_400_BAD_REQUEST)

            # Update password
            user.set_password(new_password)
            user.save()

            # Log the action
            AuditLog.log(
                action='PASSWORD_RESET',
                user=user,
                description=f'Password reset completed for {user.email}',
                request=request
            )

            # Delete the OTP record
            otp_record.delete()

            return Response({
                'message': 'Password reset successful. You can now login with your new password.'
            }, status=status.HTTP_200_OK)

        except EmailOTP.DoesNotExist:
            return Response({
                'error': 'Invalid or expired reset token. Please request a new OTP.'
            }, status=status.HTTP_400_BAD_REQUEST)


# =============================================================================
# ORGANIZATION VIEWS
# =============================================================================

class OrganizationViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing organizations.
    - Platform admins can view/edit all organizations
    - Org admins can only view/edit their own organization
    """
    queryset = Organization.objects.all()
    serializer_class = OrganizationSerializer

    def get_permissions(self):
        if self.action in ['list', 'create', 'destroy']:
            return [permissions.IsAuthenticated(), IsPlatformAdmin()]
        return [permissions.IsAuthenticated()]

    def get_queryset(self):
        user = self.request.user

        # Platform admins see all
        if getattr(user, 'is_platform_admin', False):
            return Organization.objects.all()

        # Regular users see only their organization
        if user.organization:
            return Organization.objects.filter(id=user.organization.id)

        return Organization.objects.none()

    @action(detail=False, methods=['get', 'put', 'patch'])
    def current(self, request):
        """Get or update the current user's organization"""
        organization = getattr(request, 'organization', None)

        # Fallback to user's organization if request.organization is not set
        if not organization and hasattr(request.user, 'organization'):
            organization = request.user.organization

        if not organization:
            return Response(
                {'error': 'No organization found. Please contact support.'},
                status=status.HTTP_404_NOT_FOUND
            )

        if request.method == 'GET':
            serializer = self.get_serializer(organization)
            return Response(serializer.data)

        # PUT/PATCH - only admins can update
        if request.user.role not in ['ADMIN', 'PARTNER']:
            return Response(
                {'error': 'Permission denied'},
                status=status.HTTP_403_FORBIDDEN
            )

        serializer = self.get_serializer(
            organization,
            data=request.data,
            partial=(request.method == 'PATCH')
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def usage(self, request):
        """Get usage statistics for current organization"""
        organization = getattr(request, 'organization', None)

        # Fallback to user's organization if request.organization is not set
        if not organization and hasattr(request.user, 'organization'):
            organization = request.user.organization

        if not organization:
            return Response(
                {'error': 'No organization found. Please contact support.'},
                status=status.HTTP_404_NOT_FOUND
            )

        usage = PlanService.get_usage_stats(organization)

        # Add trial info if applicable
        is_expired, days_remaining = PlanService.check_trial_expiry(organization)
        usage['trial'] = {
            'is_trial': organization.status == 'TRIAL',
            'is_expired': is_expired,
            'days_remaining': days_remaining,
        }

        return Response(usage)

    @action(detail=False, methods=['get'])
    def plans(self, request):
        """Get available subscription plans"""
        organization = getattr(request, 'organization', None)

        # Fallback to user's organization if request.organization is not set
        if not organization and hasattr(request.user, 'organization'):
            organization = request.user.organization

        current_plan = organization.plan if organization else 'FREE'
        upgrade_options = PlanService.get_upgrade_options(current_plan)

        return Response({
            'current_plan': current_plan,
            'upgrade_options': upgrade_options,
            'all_plans': PlanService.get_all_plans(),
        })

    @action(detail=False, methods=['post'])
    def request_upgrade(self, request):
        """
        Submit a plan upgrade request.
        This sends a notification to the platform admin.
        """
        organization = getattr(request, 'organization', None)
        if not organization:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_400_BAD_REQUEST
            )

        requested_plan = request.data.get('requested_plan')
        message = request.data.get('message', '')
        contact_email = request.data.get('contact_email', request.user.email)
        contact_phone = request.data.get('contact_phone', '')

        if not requested_plan:
            return Response(
                {'error': 'Requested plan is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Create or update the upgrade request
        from core.models import UpgradeRequest

        upgrade_request, created = UpgradeRequest.objects.update_or_create(
            organization=organization,
            status='PENDING',
            defaults={
                'requested_plan': requested_plan,
                'current_plan': organization.plan,
                'message': message,
                'contact_email': contact_email,
                'contact_phone': contact_phone,
                'requested_by': request.user,
            }
        )

        return Response({
            'status': 'success',
            'message': 'Your upgrade request has been submitted. Our team will contact you shortly.',
            'request_id': upgrade_request.id,
        })

    @action(detail=False, methods=['get'])
    def upgrade_requests(self, request):
        """Get upgrade request status for current organization"""
        organization = getattr(request, 'organization', None)
        if not organization:
            return Response({'requests': []})

        from core.models import UpgradeRequest
        requests = UpgradeRequest.objects.filter(organization=organization).order_by('-created_at')[:5]

        return Response({
            'requests': [
                {
                    'id': r.id,
                    'requested_plan': r.requested_plan,
                    'current_plan': r.current_plan,
                    'status': r.status,
                    'created_at': r.created_at,
                    'admin_response': r.admin_response,
                }
                for r in requests
            ]
        })


# =============================================================================
# ORGANIZATION EMAIL VIEWS
# =============================================================================

class OrganizationEmailViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing organization email accounts.
    Only admins can manage email accounts.
    """
    queryset = OrganizationEmail.objects.all()
    serializer_class = OrganizationEmailSerializer
    permission_classes = [permissions.IsAuthenticated, IsOrganizationAdmin]

    def get_queryset(self):
        """Filter by current organization"""
        organization = getattr(self.request, 'organization', None)
        if organization:
            return OrganizationEmail.objects.filter(organization=organization)
        return OrganizationEmail.objects.none()

    def get_serializer_class(self):
        if self.action in ['create', 'update', 'partial_update']:
            return OrganizationEmailWriteSerializer
        return OrganizationEmailSerializer

    def perform_create(self, serializer):
        """Set organization on create"""
        organization = getattr(self.request, 'organization', None)
        serializer.save(organization=organization)

    @action(detail=True, methods=['post'])
    def set_default(self, request, pk=None):
        """Set this email as the default for the organization"""
        email = self.get_object()
        email.is_default = True
        email.save()  # This will unset other defaults via the model's save method
        return Response({'status': 'success', 'message': 'Email set as default'})

    @action(detail=True, methods=['post'])
    def test(self, request, pk=None):
        """Send a test email using this email configuration"""
        email_account = self.get_object()
        recipient = request.data.get('recipient', request.user.email)

        if not recipient:
            return Response(
                {'status': 'error', 'message': 'Recipient email is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Use the email account's SMTP settings if configured
            success, error = EmailService.send_test_email_for_account(
                email_account=email_account,
                recipient_email=recipient
            )
            if success:
                return Response({'status': 'success', 'message': f'Test email sent to {recipient}'})
            else:
                return Response(
                    {'status': 'error', 'error': error},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {'status': 'error', 'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================================================================
# USER VIEWS
# =============================================================================

class UserViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing users within an organization"""
    queryset = User.objects.all()
    serializer_class = UserSerializer
    permission_classes = [permissions.IsAuthenticated, CanManageUsers]
    filterset_fields = ['role', 'is_active']
    search_fields = ['username', 'email', 'first_name', 'last_name']

    def get_queryset(self):
        """Get users for current organization"""
        user = self.request.user

        # Platform admins can see all users
        if getattr(user, 'is_platform_admin', False):
            org_id = self.request.headers.get('X-Organization-ID')
            if org_id:
                return User.objects.filter(organization_id=org_id)
            return User.objects.all()

        # Regular users see only their organization's users
        if hasattr(self.request, 'organization') and self.request.organization:
            return User.objects.filter(organization=self.request.organization)

        return User.objects.none()

    def perform_create(self, serializer):
        """Create user in current organization with plan limit check"""
        organization = getattr(self.request, 'organization', None)

        if organization:
            # Check user limit
            can_add, error_message = PlanService.can_add_user(organization)
            if not can_add:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'detail': error_message})

            user = serializer.save(organization=organization)

            # Send welcome email to new user and notification to admins
            from core.services.otp_service import OTPService
            OTPService.send_new_user_notification(user, organization, created_by=self.request.user)
        else:
            serializer.save()

    @action(detail=False, methods=['get', 'put'])
    def profile(self, request):
        """Get or update current user's profile"""
        user = request.user
        if request.method == 'GET':
            serializer = self.get_serializer(user)
            return Response(serializer.data)
        elif request.method == 'PUT':
            allowed_fields = ['first_name', 'last_name', 'mobile']
            update_data = {k: v for k, v in request.data.items() if k in allowed_fields}
            serializer = self.get_serializer(user, data=update_data, partial=True)
            serializer.is_valid(raise_exception=True)
            serializer.save()
            return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def change_password(self, request):
        """Change current user's password"""
        user = request.user
        current_password = request.data.get('current_password')
        new_password = request.data.get('new_password')

        if not current_password or not new_password:
            return Response(
                {'error': 'Both current_password and new_password are required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not user.check_password(current_password):
            return Response(
                {'error': 'Current password is incorrect'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if len(new_password) < 8:
            return Response(
                {'error': 'New password must be at least 8 characters long'},
                status=status.HTTP_400_BAD_REQUEST
            )

        user.set_password(new_password)
        user.save()
        return Response({'message': 'Password changed successfully'})

    @action(detail=False, methods=['get', 'put'])
    def notification_preferences(self, request):
        """Get or update current user's notification preferences"""
        user = request.user
        if request.method == 'GET':
            return Response({
                'notify_email_reminders': user.notify_email_reminders,
                'notify_task_assignments': user.notify_task_assignments,
                'notify_overdue_alerts': user.notify_overdue_alerts,
                'notify_weekly_reports': user.notify_weekly_reports,
            })
        elif request.method == 'PUT':
            allowed_fields = [
                'notify_email_reminders', 'notify_task_assignments',
                'notify_overdue_alerts', 'notify_weekly_reports'
            ]
            for field in allowed_fields:
                if field in request.data:
                    setattr(user, field, request.data[field])
            user.save()
            return Response({
                'notify_email_reminders': user.notify_email_reminders,
                'notify_task_assignments': user.notify_task_assignments,
                'notify_overdue_alerts': user.notify_overdue_alerts,
                'notify_weekly_reports': user.notify_weekly_reports,
            })

    @action(detail=False, methods=['get'])
    def export_my_data(self, request):
        """
        Export all personal data for the current user (DPDP Act compliance - Right to Data Portability).
        Returns a JSON file with all user's personal data.
        """
        user = request.user
        organization = getattr(request, 'organization', None)

        # Collect all user's personal data
        export_data = {
            'export_date': timezone.now().isoformat(),
            'export_type': 'DPDP_DATA_PORTABILITY',
            'data_controller': 'NexPro',
            'user_information': {
                'id': str(user.id),
                'username': user.username,
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'mobile': user.mobile,
                'role': user.role,
                'is_active': user.is_active,
                'date_joined': user.date_joined.isoformat() if user.date_joined else None,
                'last_login': user.last_login.isoformat() if user.last_login else None,
            },
            'notification_preferences': {
                'notify_email_reminders': user.notify_email_reminders,
                'notify_task_assignments': user.notify_task_assignments,
                'notify_overdue_alerts': user.notify_overdue_alerts,
                'notify_weekly_reports': user.notify_weekly_reports,
            },
        }

        # Add organization info if applicable
        if organization:
            export_data['organization'] = {
                'id': str(organization.id),
                'name': organization.name,
                'role_in_organization': user.role,
            }

            # Add tasks assigned to user
            from .models import WorkInstance
            user_tasks = WorkInstance.objects.filter(
                organization=organization,
                assigned_to=user
            ).values(
                'id', 'status', 'due_date', 'started_on', 'completed_on',
                'remarks', 'total_time_spent', 'created_at'
            )
            export_data['assigned_tasks'] = list(user_tasks)

        # Add audit logs for this user (last 90 days)
        from .models import AuditLog
        audit_logs = AuditLog.objects.filter(
            user=user,
            created_at__gte=timezone.now() - timezone.timedelta(days=90)
        ).values(
            'action', 'resource_type', 'details', 'ip_address', 'created_at'
        ).order_by('-created_at')[:500]  # Limit to 500 entries
        export_data['activity_logs'] = list(audit_logs)

        # Add notifications
        from .models import Notification
        notifications = Notification.objects.filter(user=user).values(
            'notification_type', 'title', 'message', 'is_read', 'created_at'
        ).order_by('-created_at')[:200]  # Limit to 200
        export_data['notifications'] = list(notifications)

        # Convert dates to strings for JSON serialization
        import json
        from django.core.serializers.json import DjangoJSONEncoder

        response = HttpResponse(
            json.dumps(export_data, cls=DjangoJSONEncoder, indent=2),
            content_type='application/json'
        )
        response['Content-Disposition'] = f'attachment; filename="nexpro_data_export_{user.username}_{timezone.now().strftime("%Y%m%d")}.json"'

        # Log the data export for audit
        AuditLog.objects.create(
            user=user,
            organization=organization,
            action='EXPORT',
            resource_type='USER_DATA',
            resource_id=str(user.id),
            details={'export_type': 'DPDP_DATA_PORTABILITY'},
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
        )

        return response

    @action(detail=False, methods=['post'])
    def request_data_deletion(self, request):
        """
        Request deletion of personal data (DPDP Act compliance - Right to Erasure).
        This initiates the data deletion process with a 30-day retention period.
        """
        user = request.user
        reason = request.data.get('reason', '')

        # Create audit log for deletion request
        from .models import AuditLog
        AuditLog.objects.create(
            user=user,
            organization=getattr(request, 'organization', None),
            action='DELETE',
            resource_type='DELETION_REQUEST',
            resource_id=str(user.id),
            details={
                'request_type': 'DPDP_RIGHT_TO_ERASURE',
                'reason': reason,
                'status': 'PENDING',
                'scheduled_deletion_date': (timezone.now() + timezone.timedelta(days=30)).isoformat()
            },
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
        )

        return Response({
            'message': 'Data deletion request received',
            'details': (
                'Your request has been logged. As per our data retention policy and DPDP Act requirements, '
                'your data will be retained for 30 days to allow you to export it or cancel the request. '
                'After this period, your personal data will be anonymized or deleted.'
            ),
            'scheduled_deletion_date': (timezone.now() + timezone.timedelta(days=30)).isoformat(),
            'contact_email': 'chinmaytechsoft@gmail.com'
        })


# =============================================================================
# CLIENT VIEWS
# =============================================================================

class ClientViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing clients"""
    queryset = Client.objects.all()
    serializer_class = ClientSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['status', 'category']
    search_fields = ['client_code', 'client_name', 'PAN', 'GSTIN', 'email']
    ordering_fields = ['client_name', 'created_at']

    def perform_create(self, serializer):
        """Create client with plan limit check"""
        organization = getattr(self.request, 'organization', None)

        if organization:
            # Check client limit
            can_add, error_message = PlanService.can_add_client(organization)
            if not can_add:
                from rest_framework.exceptions import ValidationError
                raise ValidationError({'detail': error_message})

            serializer.save(organization=organization)
        else:
            serializer.save()

    @action(detail=True, methods=['get'])
    def works(self, request, pk=None):
        """Get all work mappings for a client"""
        client = self.get_object()
        mappings = client.work_mappings.all()
        serializer = ClientWorkMappingSerializer(mappings, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def tasks(self, request, pk=None):
        """Get all work instances for a client"""
        client = self.get_object()
        instances = WorkInstance.objects.filter(client_work__client=client)
        serializer = WorkInstanceSerializer(instances, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def assign_work_types(self, request, pk=None):
        """Assign multiple task categories to a client and create initial work instances"""
        client = self.get_object()
        work_type_ids = request.data.get('work_type_ids', [])
        start_from_period = request.data.get('start_from_period', '')
        organization = getattr(request, 'organization', None)

        if not work_type_ids:
            return Response(
                {'error': 'work_type_ids is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        if not start_from_period:
            return Response(
                {'error': 'start_from_period is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        created_mappings = []
        errors = []

        for work_type_id in work_type_ids:
            try:
                # Get task category within same organization
                work_type_qs = WorkType.objects.filter(id=work_type_id)
                if organization:
                    work_type_qs = work_type_qs.filter(organization=organization)
                work_type = work_type_qs.first()

                if not work_type:
                    errors.append(f'Task category with ID {work_type_id} not found')
                    continue

                # Check if mapping already exists
                existing_mapping = ClientWorkMapping.objects.filter(
                    client=client,
                    work_type=work_type
                ).first()

                if existing_mapping:
                    if not existing_mapping.active:
                        existing_mapping.active = True
                        existing_mapping.save()
                        created_mappings.append(existing_mapping)
                    else:
                        errors.append(f'Task category {work_type.work_name} already assigned')
                    continue

                # Create new mapping
                client_work = ClientWorkMapping.objects.create(
                    client=client,
                    work_type=work_type,
                    start_from_period=start_from_period,
                    active=True,
                    organization=organization
                )

                # Auto-create first work instance
                TaskAutomationService.create_work_instance(client_work)
                created_mappings.append(client_work)

            except Exception as e:
                errors.append(f'Error assigning task category {work_type_id}: {str(e)}')

        response_data = {
            'created_count': len(created_mappings),
            'errors': errors if errors else None,
        }

        if created_mappings:
            response_data['message'] = f'Successfully assigned {len(created_mappings)} task category(s)'
            return Response(response_data, status=status.HTTP_201_CREATED)
        else:
            return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """Download Excel template for bulk client upload"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients Template"

        headers = [
            'client_code', 'client_name', 'PAN', 'GSTIN', 'email', 'mobile',
            'category', 'group', 'date_of_birth', 'date_of_incorporation',
            'status', 'address'
        ]

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        example_data = [
            'CLI001', 'Example Client Pvt Ltd', 'ABCDE1234F', '29ABCDE1234F1Z5',
            'client@example.com', '9876543210', 'COMPANY', 'Group A',
            '', '2020-01-15', 'ACTIVE', '123 Main Street, City, State - 400001'
        ]
        for col_num, value in enumerate(example_data, 1):
            ws.cell(row=2, column=col_num, value=value)

        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Bulk Client Upload Instructions"],
            [""],
            ["Required Fields:", "client_code, client_name, email, category, status"],
            [""],
            ["Category Options:", "INDIVIDUAL, FIRM, COMPANY, TRUST, HUF, OTHERS"],
            ["Status Options:", "ACTIVE, INACTIVE"],
            [""],
            ["Date Format:", "YYYY-MM-DD (e.g., 2024-01-15)"],
        ]

        for row_num, instruction in enumerate(instructions, 1):
            for col_num, text in enumerate(instruction, 1):
                cell = ws_instructions.cell(row=row_num, column=col_num, value=text)
                if row_num == 1:
                    cell.font = Font(bold=True, size=14)

        for sheet in [wb["Clients Template"], ws_instructions]:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                sheet.column_dimensions[column_letter].width = max_length + 2

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename=clients_upload_template.xlsx'
        return response

    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """Bulk upload clients from Excel file"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        organization = getattr(request, 'organization', None)

        # Check client limit before upload
        if organization:
            can_add, error_message = PlanService.can_add_client(organization)
            if not can_add:
                return Response(
                    {'error': error_message},
                    status=status.HTTP_400_BAD_REQUEST
                )

        excel_file = request.FILES['file']

        try:
            wb = load_workbook(excel_file)
            ws = wb.active

            created_clients = []
            errors = []

            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not row[0]:
                    continue

                # Check limit for each client
                if organization:
                    current_count = organization.clients.count() + len(created_clients)
                    if organization.max_clients != -1 and current_count >= organization.max_clients:
                        errors.append(f'Row {row_num}: Client limit reached')
                        break

                try:
                    client_data = {
                        'client_code': row[0],
                        'client_name': row[1],
                        'PAN': row[2] if row[2] else None,
                        'GSTIN': row[3] if row[3] else None,
                        'email': row[4],
                        'mobile': row[5] if row[5] else None,
                        'category': row[6],
                        'group': row[7] if row[7] else None,
                        'date_of_birth': row[8] if row[8] else None,
                        'date_of_incorporation': row[9] if row[9] else None,
                        'status': row[10] if row[10] else 'ACTIVE',
                        'address': row[11] if row[11] else None,
                        'organization': organization,
                    }

                    if not all([client_data['client_code'], client_data['client_name'],
                               client_data['email'], client_data['category']]):
                        errors.append(f'Row {row_num}: Missing required fields')
                        continue

                    # Check unique within organization
                    if Client.objects.filter(
                        client_code=client_data['client_code'],
                        organization=organization
                    ).exists():
                        errors.append(f'Row {row_num}: Client code already exists')
                        continue

                    client = Client.objects.create(**client_data)
                    created_clients.append(client)

                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')

            response_data = {
                'created_count': len(created_clients),
                'error_count': len(errors),
                'errors': errors if errors else None,
            }

            if created_clients:
                response_data['message'] = f'Successfully created {len(created_clients)} client(s)'
                return Response(response_data, status=status.HTTP_201_CREATED)
            else:
                response_data['message'] = 'No clients were created'
                return Response(response_data, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            return Response(
                {'error': f'Error processing file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )


# =============================================================================
# TASK CATEGORY VIEWS
# =============================================================================

class WorkTypeViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing task categories"""
    queryset = WorkType.objects.all()
    serializer_class = WorkTypeSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]
    filterset_fields = ['default_frequency', 'is_active']
    search_fields = ['work_name', 'statutory_form']
    ordering_fields = ['work_name', 'created_at']

    @action(detail=True, methods=['get'])
    def reminder_preview(self, request, pk=None):
        """Get a preview of the reminder schedule for this task category"""
        from .services.task_service import ReminderGenerationService
        from datetime import datetime

        work_type = self.get_object()

        # Get reference date from query param or use today
        ref_date_str = request.query_params.get('reference_date')
        if ref_date_str:
            try:
                reference_date = datetime.strptime(ref_date_str, '%Y-%m-%d').date()
            except ValueError:
                reference_date = None
        else:
            reference_date = None

        preview = ReminderGenerationService.get_reminder_schedule_preview(work_type, reference_date)

        # Convert dates to strings for JSON
        return Response({
            'period_start': preview['period_start'].isoformat(),
            'period_end': preview['period_end'].isoformat(),
            'due_date': preview['due_date'].isoformat(),
            'client_reminders': [d.isoformat() for d in preview['client_reminders']],
            'employee_reminders': [d.isoformat() for d in preview['employee_reminders']],
            'client_reminder_count': len(preview['client_reminders']),
            'employee_reminder_count': len(preview['employee_reminders']),
        })

    @action(detail=True, methods=['get'])
    def period_options(self, request, pk=None):
        """
        Get period label options and default due date for a task category.
        Returns relevant period options based on the task category's frequency.
        """
        from datetime import date
        from dateutil.relativedelta import relativedelta
        import calendar

        work_type = self.get_object()
        today = date.today()

        period_options = []
        default_due_date = None

        # Get period dates from task category method
        period_info = work_type.get_period_dates(today)
        default_due_date = period_info['due_date'].isoformat()

        if work_type.default_frequency == 'MONTHLY':
            # Generate monthly period options for current month and next 6 months
            for i in range(-1, 7):  # Previous month, current and next 6 months
                ref_date = today + relativedelta(months=i)
                month_name = ref_date.strftime('%b %Y')  # e.g., "Dec 2025"
                period_options.append({
                    'value': month_name,
                    'label': month_name,
                    'due_date': self._calculate_monthly_due_date(
                        ref_date.year, ref_date.month, work_type.due_date_day
                    )
                })

        elif work_type.default_frequency == 'QUARTERLY':
            # Generate quarterly period options for current FY and next FY
            # Indian FY quarters: Q1(Apr-Jun), Q2(Jul-Sep), Q3(Oct-Dec), Q4(Jan-Mar)
            if today.month >= 4:
                fy_start_year = today.year
            else:
                fy_start_year = today.year - 1

            # Generate quarters for current FY and next FY
            for fy_offset in range(0, 2):  # Current FY and next FY
                fy_year = fy_start_year + fy_offset
                fy_label = f"FY {fy_year}-{str(fy_year + 1)[-2:]}"

                quarters = [
                    ('Q1', 4, 6, f"Q1 {fy_label} (Apr-Jun)"),
                    ('Q2', 7, 9, f"Q2 {fy_label} (Jul-Sep)"),
                    ('Q3', 10, 12, f"Q3 {fy_label} (Oct-Dec)"),
                    ('Q4', 1, 3, f"Q4 {fy_label} (Jan-Mar)"),
                ]

                for q_code, start_month, end_month, q_label in quarters:
                    # Calculate due date for this quarter
                    q_year = fy_year if start_month >= 4 else fy_year + 1
                    due_month = end_month + 1
                    due_year = q_year
                    if due_month > 12:
                        due_month = 1
                        due_year += 1

                    due_day = min(work_type.due_date_day, calendar.monthrange(due_year, due_month)[1])
                    due_date = date(due_year, due_month, due_day)

                    period_options.append({
                        'value': f"{q_code} {fy_label}",
                        'label': q_label,
                        'due_date': due_date.isoformat()
                    })

        elif work_type.default_frequency == 'YEARLY':
            # Generate yearly period options (Financial Years)
            if today.month >= 4:
                current_fy_start = today.year
            else:
                current_fy_start = today.year - 1

            # Generate current FY and next 2 FYs
            for fy_offset in range(-1, 3):  # Previous, current, next 2
                fy_year = current_fy_start + fy_offset
                fy_label = f"FY {fy_year}-{str(fy_year + 1)[-2:]}"

                # Due date is typically in July of the next year for yearly tasks
                due_year = fy_year + 1
                due_month = 7  # July
                due_day = min(work_type.due_date_day, calendar.monthrange(due_year, due_month)[1])
                due_date = date(due_year, due_month, due_day)

                period_options.append({
                    'value': fy_label,
                    'label': fy_label,
                    'due_date': due_date.isoformat()
                })

        else:  # ONE_TIME
            # For one-time, just provide a generic label
            period_options.append({
                'value': f"One-time - {today.strftime('%b %Y')}",
                'label': f"One-time Task - {today.strftime('%b %Y')}",
                'due_date': (today + relativedelta(days=work_type.due_date_day)).isoformat()
            })

        return Response({
            'work_type_id': work_type.id,
            'work_type_name': work_type.work_name,
            'frequency': work_type.default_frequency,
            'due_date_day': work_type.due_date_day,
            'default_due_date': default_due_date,
            'period_options': period_options
        })

    def _calculate_monthly_due_date(self, year, month, due_day):
        """Helper to calculate due date for monthly frequency"""
        from datetime import date
        import calendar
        last_day = calendar.monthrange(year, month)[1]
        actual_due_day = min(due_day, last_day)
        return date(year, month, actual_due_day).isoformat()


# =============================================================================
# SUBTASK CATEGORY VIEWS
# =============================================================================

class SubTaskCategoryViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing subtask categories within task categories"""
    queryset = SubTaskCategory.objects.all()
    serializer_class = SubTaskCategorySerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]
    filterset_fields = ['work_type', 'is_active', 'is_required']
    search_fields = ['name', 'description']
    ordering_fields = ['order', 'name', 'created_at']

    def get_queryset(self):
        """Filter subtasks by task category if specified"""
        queryset = super().get_queryset()
        work_type_id = self.request.query_params.get('work_type')
        if work_type_id:
            queryset = queryset.filter(work_type_id=work_type_id)
        return queryset.select_related('work_type')

    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Create multiple subtask categories at once"""
        work_type_id = request.data.get('work_type')
        subtasks_data = request.data.get('subtasks', [])

        if not work_type_id:
            return Response({'error': 'work_type is required'}, status=status.HTTP_400_BAD_REQUEST)

        # Verify task category exists and belongs to organization
        try:
            work_type = WorkType.objects.get(
                id=work_type_id,
                organization=request.organization
            )
        except WorkType.DoesNotExist:
            return Response({'error': 'Task category not found'}, status=status.HTTP_404_NOT_FOUND)

        created_subtasks = []
        for idx, subtask_data in enumerate(subtasks_data):
            subtask_data['work_type'] = work_type_id
            subtask_data['organization'] = request.organization.id
            if 'order' not in subtask_data:
                subtask_data['order'] = idx

            serializer = SubTaskCategorySerializer(data=subtask_data, context={'request': request})
            if serializer.is_valid():
                subtask = serializer.save(organization=request.organization)
                created_subtasks.append(SubTaskCategorySerializer(subtask).data)
            else:
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Update task category to enable subtasks
        work_type.has_subtasks = True
        work_type.save(update_fields=['has_subtasks'])

        return Response({
            'created': len(created_subtasks),
            'subtasks': created_subtasks
        }, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def reorder(self, request):
        """Reorder subtask categories"""
        subtask_orders = request.data.get('orders', [])

        for order_data in subtask_orders:
            subtask_id = order_data.get('id')
            new_order = order_data.get('order')
            if subtask_id and new_order is not None:
                SubTaskCategory.objects.filter(
                    id=subtask_id,
                    organization=request.organization
                ).update(order=new_order)

        return Response({'status': 'reordered'})

    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate a subtask category"""
        subtask = self.get_object()

        # Create a copy with modified name
        new_subtask = SubTaskCategory.objects.create(
            organization=subtask.organization,
            work_type=subtask.work_type,
            name=f"{subtask.name} (Copy)",
            description=subtask.description,
            order=subtask.order + 1,
            is_active=subtask.is_active,
            is_required=subtask.is_required,
            is_auto_driven=subtask.is_auto_driven,
            due_days_before_parent=subtask.due_days_before_parent,
            enable_client_reminders=subtask.enable_client_reminders,
            client_reminder_start_day=subtask.client_reminder_start_day,
            client_reminder_end_day=subtask.client_reminder_end_day,
            client_reminder_frequency_type=subtask.client_reminder_frequency_type,
            client_reminder_interval_days=subtask.client_reminder_interval_days,
            client_reminder_weekdays=subtask.client_reminder_weekdays,
            enable_employee_reminders=subtask.enable_employee_reminders,
            employee_notification_type=subtask.employee_notification_type,
            employee_reminder_start_day=subtask.employee_reminder_start_day,
            employee_reminder_end_day=subtask.employee_reminder_end_day,
            employee_reminder_frequency_type=subtask.employee_reminder_frequency_type,
            employee_reminder_interval_days=subtask.employee_reminder_interval_days,
            employee_reminder_weekdays=subtask.employee_reminder_weekdays,
        )

        return Response(SubTaskCategorySerializer(new_subtask).data, status=status.HTTP_201_CREATED)


# =============================================================================
# CLIENT WORK MAPPING VIEWS
# =============================================================================

class ClientWorkMappingViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing client-work mappings"""
    queryset = ClientWorkMapping.objects.all()
    serializer_class = ClientWorkMappingSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['client', 'work_type', 'active']
    search_fields = ['client__client_name', 'work_type__work_name']

    def create(self, request, *args, **kwargs):
        """Create client work mapping and automatically create first work instance"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        organization = getattr(request, 'organization', None)
        client_work = serializer.save(organization=organization)

        # Auto-create first work instance
        TaskAutomationService.create_work_instance(client_work)

        headers = self.get_success_headers(serializer.data)
        return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)


# =============================================================================
# WORK INSTANCE VIEWS
# =============================================================================

class WorkInstanceViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing work instances (tasks)"""
    queryset = WorkInstance.objects.all()
    serializer_class = WorkInstanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['status', 'assigned_to', 'client_work__client', 'client_work__work_type']
    search_fields = ['client_work__client__client_name', 'client_work__work_type__work_name', 'period_label']
    ordering_fields = ['due_date', 'created_at']

    def list(self, request, *args, **kwargs):
        """List work instances with automatic overdue status check"""
        # Check and update overdue tasks before listing
        TaskAutomationService.check_and_update_overdue_status()
        return super().list(request, *args, **kwargs)

    def get_queryset(self):
        """Filter queryset based on user role"""
        user = self.request.user

        # Platform admins see all
        if getattr(user, 'is_platform_admin', False):
            org_id = self.request.headers.get('X-Organization-ID')
            if org_id:
                return WorkInstance.objects.filter(organization_id=org_id).select_related(
                    'client_work__client',
                    'client_work__work_type',
                    'assigned_to'
                )
            return WorkInstance.objects.all().select_related(
                'client_work__client',
                'client_work__work_type',
                'assigned_to'
            )

        base_qs = WorkInstance.objects.select_related(
            'client_work__client',
            'client_work__work_type',
            'assigned_to'
        )

        # Filter by organization
        if hasattr(self.request, 'organization') and self.request.organization:
            base_qs = base_qs.filter(organization=self.request.organization)
        else:
            return base_qs.none()

        # Staff can only see their assigned tasks
        if user.role == 'STAFF':
            return base_qs.filter(assigned_to=user)

        return base_qs

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Mark a work instance as completed"""
        work_instance = self.get_object()
        TaskAutomationService.complete_work_instance(work_instance)
        serializer = self.get_serializer(work_instance)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def update_due_date(self, request, pk=None):
        """Update due date and regenerate reminders"""
        work_instance = self.get_object()
        new_due_date = request.data.get('due_date')

        if not new_due_date:
            return Response(
                {'error': 'due_date is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        TaskAutomationService.update_due_date(work_instance, new_due_date)
        serializer = self.get_serializer(work_instance)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def reminders(self, request, pk=None):
        """Get all reminders for a work instance"""
        work_instance = self.get_object()
        reminders = work_instance.reminders.all()
        serializer = ReminderInstanceSerializer(reminders, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def transfer(self, request, pk=None):
        """Transfer a task to another employee"""
        work_instance = self.get_object()
        new_assignee_id = request.data.get('assigned_to')
        notes = request.data.get('notes', '')

        if not new_assignee_id:
            return Response(
                {'error': 'assigned_to is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            new_assignee = User.objects.get(
                id=new_assignee_id,
                organization=self.request.organization,
                is_active=True
            )
        except User.DoesNotExist:
            return Response(
                {'error': 'Invalid employee'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Check if user has permission to transfer
        # Staff can only transfer their own tasks
        if request.user.role == 'STAFF' and work_instance.assigned_to != request.user:
            return Response(
                {'error': 'You can only transfer your own tasks'},
                status=status.HTTP_403_FORBIDDEN
            )

        old_assignee = work_instance.assigned_to
        work_instance.assigned_to = new_assignee
        work_instance.save()

        # Create notification for the new assignee if they have task assignments enabled
        if new_assignee.notify_task_assignments:
            Notification.objects.create(
                organization=work_instance.organization,
                user=new_assignee,
                notification_type='TASK_ASSIGNED',
                title=f"Task Transferred: {work_instance.client_work.work_type.work_name}",
                message=f"Task '{work_instance.client_work.work_type.work_name}' for {work_instance.client_work.client.client_name} ({work_instance.period_label}) has been transferred to you by {request.user.get_full_name() or request.user.email}. {notes}",
                priority='MEDIUM',
                work_instance=work_instance,
                action_url=f"/tasks?work_instance={work_instance.id}"
            )

        serializer = self.get_serializer(work_instance)
        return Response({
            'success': True,
            'message': f"Task transferred from {old_assignee.get_full_name() if old_assignee else 'Unassigned'} to {new_assignee.get_full_name() or new_assignee.email}",
            'data': serializer.data
        })

    def update(self, request, *args, **kwargs):
        """Update work instance with automatic timer control based on status changes"""
        instance = self.get_object()
        old_status = instance.status
        new_status = request.data.get('status', old_status)

        # Call the parent update
        response = super().update(request, *args, **kwargs)

        # Handle timer based on status change
        if old_status != new_status:
            instance.refresh_from_db()
            if new_status == 'STARTED':
                # Start the timer when status changes to STARTED
                instance.start_timer()
            elif new_status in ['PAUSED', 'COMPLETED']:
                # Pause the timer when status changes to PAUSED or COMPLETED
                instance.pause_timer()

        # Re-serialize to get updated time tracking fields
        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def start_timer(self, request, pk=None):
        """Manually start the timer for a task"""
        work_instance = self.get_object()

        # Check if user has permission (staff can only control their own tasks)
        if request.user.role == 'STAFF' and work_instance.assigned_to != request.user:
            return Response(
                {'error': 'You can only control timer for your assigned tasks'},
                status=status.HTTP_403_FORBIDDEN
            )

        if work_instance.status == 'COMPLETED':
            return Response(
                {'error': 'Cannot start timer for completed task'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Update status to STARTED if not already
        if work_instance.status not in ['STARTED']:
            work_instance.status = 'STARTED'
            work_instance.save(update_fields=['status'])

        work_instance.start_timer()
        serializer = self.get_serializer(work_instance)
        return Response({
            'success': True,
            'message': 'Timer started',
            'data': serializer.data
        })

    @action(detail=True, methods=['post'])
    def pause_timer(self, request, pk=None):
        """Manually pause the timer for a task"""
        work_instance = self.get_object()

        # Check if user has permission
        if request.user.role == 'STAFF' and work_instance.assigned_to != request.user:
            return Response(
                {'error': 'You can only control timer for your assigned tasks'},
                status=status.HTTP_403_FORBIDDEN
            )

        # Update status to PAUSED
        work_instance.status = 'PAUSED'
        work_instance.save(update_fields=['status'])
        work_instance.pause_timer()

        serializer = self.get_serializer(work_instance)
        return Response({
            'success': True,
            'message': 'Timer paused',
            'time_spent': work_instance.format_time_spent(),
            'data': serializer.data
        })

    @action(detail=True, methods=['get'])
    def timer_status(self, request, pk=None):
        """Get current timer status for a task"""
        work_instance = self.get_object()
        return Response({
            'is_timer_running': work_instance.is_timer_running,
            'current_time_spent': work_instance.get_current_time_spent(),
            'formatted_time_spent': work_instance.format_time_spent(),
            'total_time_spent': work_instance.total_time_spent,
            'timer_started_at': work_instance.timer_started_at.isoformat() if work_instance.timer_started_at else None,
            'status': work_instance.status
        })

    @action(detail=True, methods=['get', 'post'])
    def documents(self, request, pk=None):
        """Get all documents for a task or upload a new document"""
        work_instance = self.get_object()

        if request.method == 'GET':
            documents = TaskDocument.objects.filter(work_instance=work_instance)
            serializer = TaskDocumentSerializer(documents, many=True, context={'request': request})
            return Response(serializer.data)

        elif request.method == 'POST':
            if 'file' not in request.FILES:
                return Response(
                    {'error': 'No file provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            data = {
                'work_instance': work_instance.id,
                'file': request.FILES['file'],
                'description': request.data.get('description', '')
            }

            serializer = TaskDocumentSerializer(data=data, context={'request': request})
            if serializer.is_valid():
                serializer.save()
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# =============================================================================
# TASK DOCUMENT VIEWS
# =============================================================================

class TaskDocumentViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing task documents"""
    queryset = TaskDocument.objects.select_related('work_instance', 'uploaded_by').all()
    serializer_class = TaskDocumentSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['work_instance']

    def get_queryset(self):
        """Filter documents based on user's access to tasks"""
        user = self.request.user

        # Platform admins see all
        if getattr(user, 'is_platform_admin', False):
            org_id = self.request.headers.get('X-Organization-ID')
            if org_id:
                return TaskDocument.objects.filter(organization_id=org_id).select_related(
                    'work_instance', 'uploaded_by'
                )
            return TaskDocument.objects.all().select_related('work_instance', 'uploaded_by')

        base_qs = TaskDocument.objects.select_related('work_instance', 'uploaded_by')

        # Filter by organization
        if hasattr(self.request, 'organization') and self.request.organization:
            base_qs = base_qs.filter(organization=self.request.organization)
        else:
            return base_qs.none()

        # Staff can only see documents for their assigned tasks
        if user.role == 'STAFF':
            return base_qs.filter(work_instance__assigned_to=user)

        return base_qs

    def create(self, request, *args, **kwargs):
        """Upload a new document"""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )

        work_instance_id = request.data.get('work_instance')
        if not work_instance_id:
            return Response(
                {'error': 'work_instance is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Verify user has access to this task
        try:
            work_instance = WorkInstance.objects.get(
                id=work_instance_id,
                organization=request.organization
            )
            # Staff can only upload to their assigned tasks
            if request.user.role == 'STAFF' and work_instance.assigned_to != request.user:
                return Response(
                    {'error': 'You can only upload documents to your assigned tasks'},
                    status=status.HTTP_403_FORBIDDEN
                )
        except WorkInstance.DoesNotExist:
            return Response(
                {'error': 'Invalid task'},
                status=status.HTTP_400_BAD_REQUEST
            )

        return super().create(request, *args, **kwargs)

    def destroy(self, request, *args, **kwargs):
        """Delete a document - only uploader or admin can delete"""
        document = self.get_object()

        # Check permissions
        if request.user.role == 'STAFF' and document.uploaded_by != request.user:
            return Response(
                {'error': 'You can only delete your own documents'},
                status=status.HTTP_403_FORBIDDEN
            )

        return super().destroy(request, *args, **kwargs)

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download a document file"""
        document = self.get_object()

        if not document.file:
            return Response(
                {'error': 'No file available'},
                status=status.HTTP_404_NOT_FOUND
            )

        from django.http import FileResponse
        import os

        file_path = document.file.path
        if os.path.exists(file_path):
            response = FileResponse(
                open(file_path, 'rb'),
                content_type=document.file_type or 'application/octet-stream'
            )
            response['Content-Disposition'] = f'attachment; filename="{document.file_name}"'
            return response

        return Response(
            {'error': 'File not found'},
            status=status.HTTP_404_NOT_FOUND
        )


# =============================================================================
# TASK CATEGORY ASSIGNMENT VIEWS
# =============================================================================

class WorkTypeAssignmentViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing task category to employee assignments"""
    queryset = WorkTypeAssignment.objects.select_related('work_type', 'employee', 'assigned_by').all()
    serializer_class = WorkTypeAssignmentSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]
    filterset_fields = ['work_type', 'employee', 'is_active']

    @action(detail=False, methods=['get'])
    def by_employee(self, request):
        """Get all task category assignments grouped by employee"""
        employee_id = request.query_params.get('employee_id')
        qs = self.get_queryset().filter(is_active=True)

        if employee_id:
            qs = qs.filter(employee_id=employee_id)

        assignments = qs.values(
            'employee__id', 'employee__email',
            'employee__first_name', 'employee__last_name'
        ).annotate(
            work_type_count=Count('id')
        ).order_by('employee__first_name')

        return Response(assignments)

    @action(detail=False, methods=['get'])
    def by_work_type(self, request):
        """Get all assignments for a specific task category"""
        work_type_id = request.query_params.get('work_type_id')
        if not work_type_id:
            return Response(
                {'error': 'work_type_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        assignments = self.get_queryset().filter(
            work_type_id=work_type_id,
            is_active=True
        )
        serializer = self.get_serializer(assignments, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def bulk_assign(self, request):
        """Assign multiple task categories to an employee"""
        employee_id = request.data.get('employee_id')
        work_type_ids = request.data.get('work_type_ids', [])

        if not employee_id:
            return Response(
                {'error': 'employee_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            employee = User.objects.get(
                id=employee_id,
                organization=self.request.organization,
                is_active=True
            )
        except User.DoesNotExist:
            return Response(
                {'error': 'Invalid employee'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get current assignments for this employee
        current_assignments = set(
            WorkTypeAssignment.objects.filter(
                employee=employee,
                organization=self.request.organization,
                is_active=True
            ).values_list('work_type_id', flat=True)
        )

        new_work_type_ids = set(work_type_ids)
        created = []
        deactivated = []

        # Deactivate assignments not in the new list
        to_deactivate = current_assignments - new_work_type_ids
        if to_deactivate:
            WorkTypeAssignment.objects.filter(
                employee=employee,
                organization=self.request.organization,
                work_type_id__in=to_deactivate
            ).update(is_active=False)
            deactivated = list(to_deactivate)

        # Create or reactivate assignments
        for wt_id in new_work_type_ids:
            assignment, was_created = WorkTypeAssignment.objects.update_or_create(
                employee=employee,
                work_type_id=wt_id,
                organization=self.request.organization,
                defaults={
                    'is_active': True,
                    'assigned_by': request.user
                }
            )
            if was_created or wt_id not in current_assignments:
                created.append(wt_id)

        return Response({
            'success': True,
            'created': len(created),
            'deactivated': len(deactivated),
            'total_active': len(new_work_type_ids)
        })


# =============================================================================
# EMAIL TEMPLATE VIEWS
# =============================================================================

class EmailTemplateViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing email templates"""
    queryset = EmailTemplate.objects.select_related('work_type').all()
    serializer_class = EmailTemplateSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]
    filterset_fields = ['work_type', 'is_active']
    search_fields = ['template_name', 'subject_template']


# =============================================================================
# REMINDER VIEWS
# =============================================================================

class ReminderRuleViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing reminder rules"""
    queryset = ReminderRule.objects.select_related('work_type', 'email_template').all()
    serializer_class = ReminderRuleSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]
    filterset_fields = ['work_type', 'reminder_type', 'is_active']


class ReminderInstanceViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing reminder instances"""
    queryset = ReminderInstance.objects.select_related(
        'work_instance__client_work__client',
        'work_instance__client_work__work_type',
        'reminder_rule'
    ).all()
    serializer_class = ReminderInstanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['send_status', 'work_instance']
    ordering_fields = ['scheduled_at', 'sent_at']


# =============================================================================
# NOTIFICATION VIEWS
# =============================================================================

class NotificationViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing user notifications"""
    serializer_class = NotificationSerializer
    permission_classes = [permissions.IsAuthenticated]
    filterset_fields = ['notification_type', 'is_read', 'priority']
    ordering_fields = ['created_at', 'is_read']

    def get_queryset(self):
        """Filter notifications for current user only"""
        user = self.request.user
        organization = getattr(self.request, 'organization', None)

        if not organization:
            return Notification.objects.none()

        return Notification.objects.filter(
            organization=organization,
            user=user
        ).select_related('work_instance').order_by('-created_at')

    @action(detail=False, methods=['get'])
    def unread_count(self, request):
        """Get count of unread notifications"""
        count = self.get_queryset().filter(is_read=False).count()
        return Response({'unread_count': count})

    @action(detail=False, methods=['post'])
    def mark_all_read(self, request):
        """Mark all notifications as read"""
        updated = self.get_queryset().filter(is_read=False).update(
            is_read=True,
            read_at=timezone.now()
        )
        return Response({'marked_read': updated})

    @action(detail=True, methods=['post'])
    def mark_read(self, request, pk=None):
        """Mark a single notification as read"""
        notification = self.get_object()
        notification.mark_as_read()
        return Response({'status': 'marked as read'})

    @action(detail=False, methods=['get'])
    def recent(self, request):
        """Get recent notifications (last 10)"""
        notifications = self.get_queryset()[:10]
        serializer = self.get_serializer(notifications, many=True)
        unread_count = self.get_queryset().filter(is_read=False).count()
        return Response({
            'notifications': serializer.data,
            'unread_count': unread_count
        })


# =============================================================================
# CREDENTIAL VAULT VIEWS
# =============================================================================

class CredentialVaultViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing client portal credentials (encrypted)"""
    serializer_class = CredentialVaultSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrPartner]
    filterset_fields = ['client', 'portal_type']
    ordering_fields = ['client__client_name', 'portal_type', 'last_updated']

    def get_queryset(self):
        """Filter credentials by organization"""
        organization = getattr(self.request, 'organization', None)
        if not organization:
            return CredentialVault.objects.none()

        return CredentialVault.objects.filter(
            organization=organization
        ).select_related('client').order_by('client__client_name', 'portal_type')

    @action(detail=False, methods=['get'])
    def by_client(self, request):
        """Get all credentials for a specific client"""
        client_id = request.query_params.get('client_id')
        if not client_id:
            return Response({'error': 'client_id is required'}, status=400)

        credentials = self.get_queryset().filter(client_id=client_id)
        serializer = self.get_serializer(credentials, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def reveal(self, request, pk=None):
        """Reveal (decrypt) the password for a credential - with audit logging"""
        credential = self.get_object()
        try:
            decrypted_password = credential.decrypt_password()

            # Log the access for audit purposes
            AuditLog.objects.create(
                organization=request.organization,
                user=request.user,
                action='CREDENTIAL_ACCESS',
                resource_type='CredentialVault',
                resource_id=str(credential.id),
                details={
                    'client_id': str(credential.client.id),
                    'client_name': credential.client.client_name,
                    'portal_type': credential.portal_type
                },
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:500]
            )

            return Response({
                'password': decrypted_password,
                'portal_type': credential.portal_type,
                'username': credential.username
            })
        except Exception as e:
            return Response({'error': 'Failed to decrypt password'}, status=500)

    @action(detail=False, methods=['get'])
    def portal_types(self, request):
        """Get available portal types"""
        return Response([
            {'value': choice[0], 'label': choice[1]}
            for choice in CredentialVault.PORTAL_TYPE_CHOICES
        ])


# =============================================================================
# DASHBOARD VIEWS
# =============================================================================

class DashboardViewSet(viewsets.ViewSet):
    """Dashboard statistics and summary"""
    permission_classes = [permissions.IsAuthenticated]

    def _get_base_queryset(self, request):
        """Get base queryset filtered by organization and user role"""
        user = request.user
        organization = getattr(request, 'organization', None)

        if not organization:
            return WorkInstance.objects.none()

        base_qs = WorkInstance.objects.filter(organization=organization)

        # Staff can only see their assigned tasks
        if user.role == 'STAFF':
            return base_qs.filter(assigned_to=user)

        return base_qs

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get dashboard summary stats"""
        today = timezone.now().date()
        week_later = today + timedelta(days=7)
        user = request.user
        organization = getattr(request, 'organization', None)

        base_qs = self._get_base_queryset(request)

        if user.role == 'STAFF':
            stats = {
                'my_tasks': base_qs.filter(
                    status__in=['NOT_STARTED', 'STARTED', 'IN_PROGRESS']
                ).count(),
                'pending_tasks': base_qs.filter(
                    status__in=['NOT_STARTED', 'STARTED', 'IN_PROGRESS']
                ).count(),
                'overdue_tasks': base_qs.filter(status='OVERDUE').count(),
                'today_due': base_qs.filter(due_date=today).count(),
                'week_due': base_qs.filter(
                    due_date__gte=today,
                    due_date__lte=week_later
                ).count(),
                'completed_this_month': base_qs.filter(
                    status='COMPLETED',
                    completed_on__year=today.year,
                    completed_on__month=today.month
                ).count(),
            }
        else:
            stats = {
                'total_clients': Client.objects.filter(
                    organization=organization,
                    status='ACTIVE'
                ).count() if organization else 0,
                'pending_tasks': base_qs.filter(
                    status__in=['NOT_STARTED', 'STARTED', 'IN_PROGRESS']
                ).count(),
                'overdue_tasks': base_qs.filter(status='OVERDUE').count(),
                'today_due': base_qs.filter(due_date=today).count(),
                'week_due': base_qs.filter(
                    due_date__gte=today,
                    due_date__lte=week_later
                ).count(),
                'completed_this_month': base_qs.filter(
                    status='COMPLETED',
                    completed_on__year=today.year,
                    completed_on__month=today.month
                ).count(),
            }

        return Response(stats)

    @action(detail=False, methods=['get'])
    def upcoming_tasks(self, request):
        """Get upcoming tasks for dashboard"""
        limit = int(request.query_params.get('limit', 10))
        base_qs = self._get_base_queryset(request)

        tasks = base_qs.filter(
            status__in=['NOT_STARTED', 'STARTED', 'IN_PROGRESS', 'OVERDUE']
        ).order_by('due_date')[:limit]

        serializer = WorkInstanceSerializer(tasks, many=True, context={'request': request})
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def analytics(self, request):
        """Get detailed analytics for reports"""
        organization = getattr(request, 'organization', None)

        if not organization:
            return Response({})

        base_qs = WorkInstance.objects.filter(organization=organization)

        # Status distribution
        status_counts = {}
        for status_choice in ['NOT_STARTED', 'STARTED', 'IN_PROGRESS', 'COMPLETED', 'OVERDUE']:
            status_counts[status_choice] = base_qs.filter(status=status_choice).count()

        # Work type distribution
        work_type_counts = base_qs.values(
            'client_work__work_type__work_name'
        ).annotate(count=Count('id')).order_by('-count')[:10]

        # Client-wise task distribution
        client_counts = base_qs.values(
            'client_work__client__client_name'
        ).annotate(count=Count('id')).order_by('-count')[:10]

        # Staff productivity
        staff_counts = base_qs.values(
            'assigned_to__username'
        ).annotate(
            total=Count('id'),
            completed=Count('id', filter=Q(status='COMPLETED')),
            pending=Count('id', filter=Q(status__in=['NOT_STARTED', 'STARTED', 'IN_PROGRESS'])),
            overdue=Count('id', filter=Q(status='OVERDUE'))
        ).order_by('-total')

        return Response({
            'status_distribution': status_counts,
            'work_type_distribution': list(work_type_counts),
            'client_distribution': list(client_counts),
            'staff_productivity': list(staff_counts),
        })

    @action(detail=False, methods=['get'])
    def email_config(self, request):
        """Get email configuration status"""
        config_status = EmailService.get_email_configuration_status()
        return Response(config_status)

    @action(detail=False, methods=['post'])
    def test_email(self, request):
        """Send a test email to verify configuration"""
        recipient = request.data.get('recipient_email')

        if not recipient:
            return Response(
                {'error': 'recipient_email is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        success, error = EmailService.send_test_email(recipient)

        if success:
            return Response({'message': 'Test email sent successfully'})
        else:
            return Response(
                {'error': f'Failed to send test email: {error}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================================================================
# PLATFORM ADMIN VIEWS (SuperAdmin Dashboard)
# =============================================================================

class PlatformAdminViewSet(viewsets.ViewSet):
    """
    Platform Admin dashboard and management.
    Only accessible by platform admins (is_platform_admin=True).
    """
    permission_classes = [permissions.IsAuthenticated, IsPlatformAdmin]

    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get platform-wide statistics for SuperAdmin dashboard"""
        today = timezone.now().date()
        this_month_start = today.replace(day=1)

        # Organization stats
        total_orgs = Organization.objects.count()
        active_orgs = Organization.objects.filter(status='ACTIVE').count()
        trial_orgs = Organization.objects.filter(status='TRIAL').count()
        suspended_orgs = Organization.objects.filter(status='SUSPENDED').count()

        # Trial expiry stats
        expiring_soon = Organization.objects.filter(
            status='TRIAL',
            trial_ends_at__lte=timezone.now() + timedelta(days=7),
            trial_ends_at__gt=timezone.now()
        ).count()

        expired_trials = Organization.objects.filter(
            status='TRIAL',
            trial_ends_at__lte=timezone.now()
        ).count()

        # User stats (excluding platform admins)
        total_users = User.objects.filter(is_platform_admin=False).count()
        active_users = User.objects.filter(is_platform_admin=False, is_active=True).count()

        # New organizations this month
        new_orgs_this_month = Organization.objects.filter(
            created_at__gte=this_month_start
        ).count()

        # Plan distribution - dynamically from SubscriptionPlan model
        plan_distribution = {}
        # First, get counts from SubscriptionPlan records
        subscription_plans = SubscriptionPlan.objects.filter(is_active=True)
        for plan in subscription_plans:
            plan_distribution[plan.code] = Organization.objects.filter(plan=plan.code).count()
        # Also include any legacy/hardcoded plan codes that may still be in use
        legacy_codes = ['FREE', 'STARTER', 'PROFESSIONAL', 'ENTERPRISE']
        for plan_code in legacy_codes:
            if plan_code not in plan_distribution:
                count = Organization.objects.filter(plan=plan_code).count()
                if count > 0:  # Only include if there are orgs using this plan
                    plan_distribution[plan_code] = count

        # Total clients and tasks across platform
        total_clients = Client.objects.count()
        total_tasks = WorkInstance.objects.count()
        pending_tasks = WorkInstance.objects.filter(
            status__in=['NOT_STARTED', 'STARTED', 'IN_PROGRESS']
        ).count()

        return Response({
            'organizations': {
                'total': total_orgs,
                'active': active_orgs,
                'trial': trial_orgs,
                'suspended': suspended_orgs,
                'new_this_month': new_orgs_this_month,
            },
            'trials': {
                'expiring_soon': expiring_soon,
                'expired': expired_trials,
            },
            'users': {
                'total': total_users,
                'active': active_users,
            },
            'plan_distribution': plan_distribution,
            'platform': {
                'total_clients': total_clients,
                'total_tasks': total_tasks,
                'pending_tasks': pending_tasks,
            }
        })

    @action(detail=False, methods=['get'])
    def organizations(self, request):
        """Get list of all organizations with details"""
        orgs = Organization.objects.annotate(
            user_count_ann=Count('users', filter=Q(users__is_platform_admin=False)),
            client_count_ann=Count('client_set'),
            task_count_ann=Count('workinstance_set')
        ).order_by('-created_at')

        org_data = []
        for org in orgs:
            is_expired, days_remaining = PlanService.check_trial_expiry(org)
            org_data.append({
                'id': str(org.id),
                'name': org.name,
                'email': org.email,
                'plan': org.plan,
                'plan_name': PlanService.get_plan(org.plan)['name'],
                'status': org.status,
                'user_count': org.user_count_ann,
                'client_count': org.client_count_ann,
                'task_count': org.task_count_ann,
                'max_users': org.max_users,
                'max_clients': org.max_clients,
                'trial_ends_at': org.trial_ends_at,
                'is_trial_expired': is_expired,
                'days_remaining': days_remaining,
                'created_at': org.created_at,
            })

        return Response(org_data)

    @action(detail=False, methods=['get'])
    def recent_signups(self, request):
        """Get recent organization signups"""
        limit = int(request.query_params.get('limit', 10))

        recent_orgs = Organization.objects.order_by('-created_at')[:limit]

        org_data = []
        for org in recent_orgs:
            admin_user = org.users.filter(role='ADMIN').first()
            org_data.append({
                'id': str(org.id),
                'name': org.name,
                'email': org.email,
                'plan': org.plan,
                'status': org.status,
                'admin_name': f"{admin_user.first_name} {admin_user.last_name}".strip() if admin_user else 'N/A',
                'admin_email': admin_user.email if admin_user else 'N/A',
                'created_at': org.created_at,
            })

        return Response(org_data)

    @action(detail=False, methods=['get'])
    def expiring_trials(self, request):
        """Get organizations with expiring or expired trials"""
        expiring = Organization.objects.filter(
            status='TRIAL'
        ).order_by('trial_ends_at')

        trial_data = []
        for org in expiring:
            is_expired, days_remaining = PlanService.check_trial_expiry(org)
            trial_data.append({
                'id': str(org.id),
                'name': org.name,
                'email': org.email,
                'trial_ends_at': org.trial_ends_at,
                'is_expired': is_expired,
                'days_remaining': days_remaining,
                'user_count': org.users.filter(is_platform_admin=False).count(),
                'client_count': org.client_set.count(),
            })

        return Response(trial_data)

    @action(detail=True, methods=['post'])
    def suspend(self, request, pk=None):
        """Suspend an organization"""
        try:
            org = Organization.objects.get(id=pk)
            org.status = 'SUSPENDED'
            org.save()
            return Response({'message': f'Organization {org.name} has been suspended'})
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['post'])
    def activate(self, request, pk=None):
        """Activate a suspended organization"""
        try:
            org = Organization.objects.get(id=pk)
            org.status = 'ACTIVE'
            org.save()
            return Response({'message': f'Organization {org.name} has been activated'})
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['post'])
    def extend_trial(self, request, pk=None):
        """Extend trial period for an organization"""
        try:
            org = Organization.objects.get(id=pk)
            days = int(request.data.get('days', 14))

            if org.trial_ends_at:
                org.trial_ends_at = org.trial_ends_at + timedelta(days=days)
            else:
                org.trial_ends_at = timezone.now() + timedelta(days=days)

            org.status = 'TRIAL'
            org.save()

            return Response({
                'message': f'Trial extended by {days} days',
                'new_trial_end': org.trial_ends_at
            })
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=True, methods=['post'])
    def upgrade_plan(self, request, pk=None):
        """Upgrade organization to a new plan"""
        try:
            org = Organization.objects.get(id=pk)
            new_plan = request.data.get('plan')

            if not new_plan:
                return Response(
                    {'error': 'Plan code is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            try:
                PlanService.upgrade_plan(org, new_plan)
                return Response({
                    'message': f'Organization upgraded to {new_plan}',
                    'organization': OrganizationSerializer(org).data
                })
            except ValueError as e:
                return Response(
                    {'error': str(e)},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=False, methods=['get'])
    def all_users(self, request):
        """Get all users across all organizations"""
        users = User.objects.filter(is_platform_admin=False).select_related('organization').order_by('-date_joined')

        user_data = []
        for user in users:
            user_data.append({
                'id': str(user.id),
                'email': user.email,
                'first_name': user.first_name,
                'last_name': user.last_name,
                'role': user.role,
                'is_active': user.is_active,
                'organization_id': str(user.organization.id) if user.organization else None,
                'organization_name': user.organization.name if user.organization else 'N/A',
                'organization_status': user.organization.status if user.organization else 'N/A',
                'date_joined': user.date_joined,
                'last_login': user.last_login,
            })

        return Response(user_data)

    @action(detail=True, methods=['get'])
    def org_details(self, request, pk=None):
        """Get detailed information about an organization"""
        try:
            org = Organization.objects.get(id=pk)

            # Get usage stats
            usage = PlanService.get_usage_stats(org)
            is_expired, days_remaining = PlanService.check_trial_expiry(org)

            # Get users
            users = org.users.filter(is_platform_admin=False).values(
                'id', 'email', 'first_name', 'last_name', 'role', 'is_active'
            )

            # Get recent activity
            recent_tasks = WorkInstance.objects.filter(
                organization=org
            ).order_by('-created_at')[:5].values(
                'id', 'status', 'period_label', 'due_date',
                'client_work__client__client_name',
                'client_work__work_type__work_name'
            )

            return Response({
                'organization': OrganizationSerializer(org).data,
                'usage': usage,
                'trial': {
                    'is_trial': org.status == 'TRIAL',
                    'is_expired': is_expired,
                    'days_remaining': days_remaining,
                },
                'users': list(users),
                'recent_tasks': list(recent_tasks),
            })
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )

    @action(detail=False, methods=['get', 'put', 'patch'], url_path='settings')
    def platform_settings(self, request):
        """Get or update platform settings"""
        platform_settings_obj = PlatformSettings.get_settings()

        if request.method == 'GET':
            # Audit log for viewing settings
            AuditLogger.log_from_request(
                action=AuditAction.PLATFORM_SETTINGS_VIEW,
                request=request,
                success=True,
                resource_type='PlatformSettings',
                resource_id=platform_settings_obj.id
            )
            serializer = PlatformSettingsSerializer(platform_settings_obj)
            return Response(serializer.data)
        else:
            # PUT or PATCH
            serializer = PlatformSettingsSerializer(
                platform_settings_obj,
                data=request.data,
                partial=request.method == 'PATCH'
            )
            if serializer.is_valid():
                # Determine which settings are being changed
                changed_fields = list(request.data.keys())
                is_smtp_change = any(f.startswith('smtp_') for f in changed_fields)
                is_google_change = any(f.startswith('google_') for f in changed_fields)

                serializer.save()

                # Audit log for updating settings
                audit_action = AuditAction.PLATFORM_SETTINGS_UPDATE
                if is_smtp_change:
                    audit_action = AuditAction.SMTP_CONFIG_UPDATE
                elif is_google_change:
                    audit_action = AuditAction.GOOGLE_OAUTH_CONFIG_UPDATE

                AuditLogger.log_from_request(
                    action=audit_action,
                    request=request,
                    success=True,
                    resource_type='PlatformSettings',
                    resource_id=platform_settings_obj.id,
                    details={'changed_fields': changed_fields}
                )

                return Response(serializer.data)

            # Log failed update attempt
            AuditLogger.log_from_request(
                action=AuditAction.PLATFORM_SETTINGS_UPDATE,
                request=request,
                success=False,
                resource_type='PlatformSettings',
                resource_id=platform_settings_obj.id,
                details={'errors': serializer.errors}
            )
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'], url_path='test_smtp')
    def test_smtp(self, request):
        """Send a test email using configured SMTP settings"""
        import smtplib
        from email.mime.text import MIMEText
        from email.mime.multipart import MIMEMultipart

        platform_settings_obj = PlatformSettings.get_settings()

        if not platform_settings_obj.smtp_enabled:
            return Response(
                {'error': 'SMTP is not enabled. Enable SMTP in settings first.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        recipient_email = request.data.get('recipient_email')
        if not recipient_email:
            return Response(
                {'error': 'recipient_email is required'},
                status=status.HTTP_400_BAD_REQUEST
            )

        # Validate SMTP configuration
        if not platform_settings_obj.smtp_host:
            return Response(
                {'error': 'SMTP host is not configured'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not platform_settings_obj.smtp_username:
            return Response(
                {'error': 'SMTP username is not configured'},
                status=status.HTTP_400_BAD_REQUEST
            )
        if not platform_settings_obj.smtp_password:
            return Response(
                {'error': 'SMTP password is not configured'},
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            # Create message
            msg = MIMEMultipart('alternative')
            msg['Subject'] = f'Test Email from {platform_settings_obj.platform_name}'
            msg['From'] = f'{platform_settings_obj.smtp_from_name or platform_settings_obj.platform_name} <{platform_settings_obj.smtp_from_email or platform_settings_obj.smtp_username}>'
            msg['To'] = recipient_email

            # Email content
            text_content = f"""
This is a test email from {platform_settings_obj.platform_name}.

If you received this email, your SMTP configuration is working correctly!

SMTP Settings:
- Host: {platform_settings_obj.smtp_host}
- Port: {platform_settings_obj.smtp_port}
- TLS: {'Yes' if platform_settings_obj.smtp_use_tls else 'No'}
- SSL: {'Yes' if platform_settings_obj.smtp_use_ssl else 'No'}

Best regards,
{platform_settings_obj.platform_name} Team
            """

            html_content = f"""
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
        <h2 style="color: #667eea;">Test Email from {platform_settings_obj.platform_name}</h2>
        <p>If you received this email, your SMTP configuration is working correctly!</p>
        <div style="background: #f8f9fa; padding: 15px; border-radius: 8px; margin: 20px 0;">
            <h4 style="margin: 0 0 10px 0;">SMTP Settings:</h4>
            <ul style="margin: 0; padding-left: 20px;">
                <li><strong>Host:</strong> {platform_settings_obj.smtp_host}</li>
                <li><strong>Port:</strong> {platform_settings_obj.smtp_port}</li>
                <li><strong>TLS:</strong> {'Yes' if platform_settings_obj.smtp_use_tls else 'No'}</li>
                <li><strong>SSL:</strong> {'Yes' if platform_settings_obj.smtp_use_ssl else 'No'}</li>
            </ul>
        </div>
        <p style="color: #666;">Best regards,<br>{platform_settings_obj.platform_name} Team</p>
    </div>
</body>
</html>
            """

            msg.attach(MIMEText(text_content, 'plain'))
            msg.attach(MIMEText(html_content, 'html'))

            # Send email
            if platform_settings_obj.smtp_use_ssl:
                server = smtplib.SMTP_SSL(
                    platform_settings_obj.smtp_host,
                    platform_settings_obj.smtp_port
                )
            else:
                server = smtplib.SMTP(
                    platform_settings_obj.smtp_host,
                    platform_settings_obj.smtp_port
                )
                if platform_settings_obj.smtp_use_tls:
                    server.starttls()

            server.login(
                platform_settings_obj.smtp_username,
                platform_settings_obj.smtp_password
            )
            server.sendmail(
                platform_settings_obj.smtp_from_email or platform_settings_obj.smtp_username,
                recipient_email,
                msg.as_string()
            )
            server.quit()

            # Audit log for successful test email
            AuditLogger.log_from_request(
                action=AuditAction.SMTP_TEST_EMAIL,
                request=request,
                success=True,
                resource_type='PlatformSettings',
                details={'recipient': recipient_email}
            )

            return Response({
                'message': f'Test email sent successfully to {recipient_email}',
                'recipient': recipient_email
            })

        except smtplib.SMTPAuthenticationError as e:
            AuditLogger.log_from_request(
                action=AuditAction.SMTP_TEST_EMAIL,
                request=request,
                success=False,
                resource_type='PlatformSettings',
                details={'recipient': recipient_email, 'error': 'Authentication failed'}
            )
            return Response(
                {'error': f'SMTP Authentication failed. Please check your username and password. Error: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except smtplib.SMTPConnectError as e:
            AuditLogger.log_from_request(
                action=AuditAction.SMTP_TEST_EMAIL,
                request=request,
                success=False,
                resource_type='PlatformSettings',
                details={'recipient': recipient_email, 'error': 'Connection failed'}
            )
            return Response(
                {'error': f'Failed to connect to SMTP server. Please check host and port. Error: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        except Exception as e:
            AuditLogger.log_from_request(
                action=AuditAction.SMTP_TEST_EMAIL,
                request=request,
                success=False,
                resource_type='PlatformSettings',
                details={'recipient': recipient_email, 'error': str(e)}
            )
            return Response(
                {'error': f'Failed to send test email: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['delete'], url_path='delete_org')
    def delete_organization(self, request, pk=None):
        """
        Permanently delete an organization and all its data.
        This is a destructive action and cannot be undone.
        """
        try:
            org = Organization.objects.get(id=pk)
            org_name = org.name

            # Delete all related data (cascade should handle most, but be explicit)
            # Users associated with the organization
            User.objects.filter(organization=org, is_platform_admin=False).delete()

            # Delete the organization (cascades to clients, works, etc.)
            org.delete()

            return Response({
                'message': f'Organization "{org_name}" and all associated data have been permanently deleted'
            })
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to delete organization: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['delete'], url_path='delete_user')
    def delete_user(self, request, pk=None):
        """
        Permanently delete a user.
        Platform admins cannot be deleted through this endpoint.
        """
        try:
            user = User.objects.get(id=pk)

            if user.is_platform_admin:
                return Response(
                    {'error': 'Cannot delete platform admin users through this endpoint'},
                    status=status.HTTP_403_FORBIDDEN
                )

            user_email = user.email
            user.delete()

            return Response({
                'message': f'User "{user_email}" has been permanently deleted'
            })
        except User.DoesNotExist:
            return Response(
                {'error': 'User not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to delete user: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['post'], url_path='change_plan')
    def change_plan(self, request, pk=None):
        """
        Change organization plan (upgrade or downgrade).
        This replaces the upgrade_plan action with more flexibility.
        """
        try:
            org = Organization.objects.get(id=pk)
            new_plan = request.data.get('plan')

            if not new_plan:
                return Response(
                    {'error': 'Plan code is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            valid_plans = ['FREE', 'STARTER', 'PROFESSIONAL', 'ENTERPRISE']
            if new_plan not in valid_plans:
                return Response(
                    {'error': f'Invalid plan. Must be one of: {", ".join(valid_plans)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )

            old_plan = org.plan

            # Update plan and limits based on plan type
            plan_limits = {
                'FREE': {'max_users': 2, 'max_clients': 10},
                'STARTER': {'max_users': 5, 'max_clients': 50},
                'PROFESSIONAL': {'max_users': 15, 'max_clients': 200},
                'ENTERPRISE': {'max_users': 999, 'max_clients': 9999},
            }

            org.plan = new_plan
            org.max_users = plan_limits[new_plan]['max_users']
            org.max_clients = plan_limits[new_plan]['max_clients']

            # If upgrading from FREE/TRIAL, activate the organization
            if new_plan != 'FREE' and org.status == 'TRIAL':
                org.status = 'ACTIVE'

            org.save()

            action_type = 'upgraded' if valid_plans.index(new_plan) > valid_plans.index(old_plan) else 'downgraded'

            return Response({
                'message': f'Organization {action_type} from {old_plan} to {new_plan}',
                'organization': OrganizationSerializer(org).data
            })
        except Organization.DoesNotExist:
            return Response(
                {'error': 'Organization not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': f'Failed to change plan: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# =============================================================================
# SUBSCRIPTION PLAN VIEWS (Super Admin)
# =============================================================================

class SubscriptionPlanViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing subscription plans.
    Only accessible by platform admins.
    """
    queryset = SubscriptionPlan.objects.all()
    serializer_class = SubscriptionPlanSerializer
    permission_classes = [permissions.IsAuthenticated, IsPlatformAdmin]

    def get_queryset(self):
        """Return all plans, optionally filtered by status"""
        queryset = SubscriptionPlan.objects.all()
        is_active = self.request.query_params.get('is_active', None)
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        return queryset

    def create(self, request, *args, **kwargs):
        """Create a new subscription plan"""
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    def destroy(self, request, *args, **kwargs):
        """Delete a subscription plan"""
        instance = self.get_object()

        # Check if any organizations are using this plan
        org_count = Organization.objects.filter(plan=instance.code).count()
        if org_count > 0:
            return Response(
                {'error': f'Cannot delete plan. {org_count} organization(s) are currently using this plan.'},
                status=status.HTTP_400_BAD_REQUEST
            )

        instance.delete()
        return Response({'message': 'Plan deleted successfully'}, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'])
    def set_default(self, request, pk=None):
        """Set this plan as the default for new signups"""
        plan = self.get_object()
        plan.is_default = True
        plan.save()  # save() method handles unsetting other defaults
        return Response({
            'message': f'{plan.name} is now the default plan',
            'plan': SubscriptionPlanSerializer(plan).data
        })

    @action(detail=False, methods=['get'], permission_classes=[permissions.AllowAny])
    def public(self, request):
        """Get all active subscription plans for public display (landing page pricing)"""
        plans = SubscriptionPlan.objects.filter(is_active=True).order_by('sort_order', 'price_monthly')
        serializer = SubscriptionPlanSerializer(plans, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['post'])
    def initialize_defaults(self, request):
        """Initialize default subscription plans if none exist"""
        if SubscriptionPlan.objects.exists():
            return Response({
                'message': 'Plans already exist. No changes made.',
                'count': SubscriptionPlan.objects.count()
            })

        SubscriptionPlan.create_default_plans()
        return Response({
            'message': 'Default plans created successfully',
            'plans': SubscriptionPlanSerializer(SubscriptionPlan.objects.all(), many=True).data
        })

    @action(detail=False, methods=['get'])
    def sync_frequency_choices(self, request):
        """Get available sync frequency options for plans"""
        choices = [
            {'value': value, 'label': label}
            for value, label in SubscriptionPlan.SYNC_FREQUENCY_CHOICES
        ]
        return Response(choices)


# =============================================================================
# GOOGLE API QUOTA MONITORING (Super Admin)
# =============================================================================

class GoogleQuotaViewSet(viewsets.ViewSet):
    """
    ViewSet for monitoring Google API quota usage.
    Only accessible by platform admins.
    """
    permission_classes = [permissions.IsAuthenticated, IsPlatformAdmin]

    @action(detail=False, methods=['get'])
    def summary(self, request):
        """Get current day's quota usage summary"""
        from datetime import datetime
        date_str = request.query_params.get('date', None)

        if date_str:
            try:
                date = datetime.strptime(date_str, '%Y-%m-%d').date()
            except ValueError:
                return Response(
                    {'error': 'Invalid date format. Use YYYY-MM-DD'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            date = None

        summary = GoogleAPIQuotaUsage.get_daily_summary(date)

        # Add capacity estimates
        platform_settings = PlatformSettings.get_settings()
        tasks_usage = summary['apis'].get('TASKS', {}).get('queries', 0)
        tasks_quota = platform_settings.google_tasks_daily_quota

        # Estimate tenant capacity based on Tasks API (bottleneck)
        if tasks_usage > 0:
            avg_queries_per_tenant = tasks_usage / max(1, summary['apis'].get('TASKS', {}).get('unique_organizations', 1))
            estimated_capacity = int(tasks_quota / max(1, avg_queries_per_tenant))
        else:
            # Default estimate: hourly sync = ~50 queries/user/day
            estimated_capacity = int(tasks_quota / 50)

        summary['total_queries'] = sum(
            api.get('queries', 0) for api in summary['apis'].values()
        )
        summary['estimated_tenant_capacity'] = estimated_capacity
        summary['quota_settings'] = {
            'tasks_daily_quota': platform_settings.google_tasks_daily_quota,
            'calendar_daily_quota': platform_settings.google_calendar_daily_quota,
            'drive_daily_quota': platform_settings.google_drive_daily_quota,
            'gmail_daily_quota': platform_settings.google_gmail_daily_quota,
            'warning_threshold': platform_settings.quota_warning_threshold,
            'critical_threshold': platform_settings.quota_critical_threshold,
        }

        return Response(summary)

    @action(detail=False, methods=['get'])
    def history(self, request):
        """Get quota usage history for the past N days"""
        from datetime import timedelta
        from django.utils import timezone

        days = int(request.query_params.get('days', 7))
        days = min(days, 90)  # Max 90 days

        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)

        history = []
        current_date = start_date

        while current_date <= end_date:
            daily_summary = GoogleAPIQuotaUsage.get_daily_summary(current_date)
            history.append(daily_summary)
            current_date += timedelta(days=1)

        return Response({
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'days': days,
            'history': history
        })

    @action(detail=False, methods=['get'])
    def by_api(self, request):
        """Get detailed usage breakdown by API type"""
        from datetime import timedelta
        from django.utils import timezone
        from django.db.models import Sum

        api_type = request.query_params.get('api_type', 'TASKS')
        days = int(request.query_params.get('days', 7))

        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)

        usage_data = GoogleAPIQuotaUsage.objects.filter(
            api_type=api_type,
            date__gte=start_date,
            date__lte=end_date
        ).order_by('date')

        serializer = GoogleAPIQuotaUsageSerializer(usage_data, many=True)

        totals = usage_data.aggregate(
            total_queries=Sum('queries_count'),
            total_failed=Sum('failed_requests'),
            total_rate_limits=Sum('rate_limit_hits')
        )

        return Response({
            'api_type': api_type,
            'start_date': start_date.isoformat(),
            'end_date': end_date.isoformat(),
            'totals': totals,
            'daily_usage': serializer.data
        })

    @action(detail=False, methods=['get'])
    def alerts(self, request):
        """Get current quota alerts/warnings"""
        platform_settings = PlatformSettings.get_settings()
        summary = GoogleAPIQuotaUsage.get_daily_summary()

        alerts = []
        for api_type, data in summary['apis'].items():
            percentage = data.get('percentage', 0)
            if percentage >= platform_settings.quota_critical_threshold:
                alerts.append({
                    'level': 'critical',
                    'api': api_type,
                    'message': f'{api_type} API usage is at {percentage}% of daily quota',
                    'percentage': percentage,
                    'queries': data.get('queries', 0),
                    'quota': data.get('quota_limit', 0)
                })
            elif percentage >= platform_settings.quota_warning_threshold:
                alerts.append({
                    'level': 'warning',
                    'api': api_type,
                    'message': f'{api_type} API usage is at {percentage}% of daily quota',
                    'percentage': percentage,
                    'queries': data.get('queries', 0),
                    'quota': data.get('quota_limit', 0)
                })

        # Check for rate limit hits
        for api_type, data in summary['apis'].items():
            rate_limits = data.get('rate_limits', 0)
            if rate_limits > 0:
                alerts.append({
                    'level': 'warning',
                    'api': api_type,
                    'message': f'{api_type} API hit rate limits {rate_limits} times today',
                    'rate_limit_hits': rate_limits
                })

        return Response({
            'date': summary['date'],
            'alerts_count': len(alerts),
            'alerts': alerts
        })

    @action(detail=False, methods=['post'])
    def update_quotas(self, request):
        """Update quota limits (Super Admin only)"""
        platform_settings = PlatformSettings.get_settings()

        # Update fields if provided
        if 'google_tasks_daily_quota' in request.data:
            platform_settings.google_tasks_daily_quota = request.data['google_tasks_daily_quota']
        if 'google_calendar_daily_quota' in request.data:
            platform_settings.google_calendar_daily_quota = request.data['google_calendar_daily_quota']
        if 'google_drive_daily_quota' in request.data:
            platform_settings.google_drive_daily_quota = request.data['google_drive_daily_quota']
        if 'google_gmail_daily_quota' in request.data:
            platform_settings.google_gmail_daily_quota = request.data['google_gmail_daily_quota']
        if 'quota_warning_threshold' in request.data:
            platform_settings.quota_warning_threshold = request.data['quota_warning_threshold']
        if 'quota_critical_threshold' in request.data:
            platform_settings.quota_critical_threshold = request.data['quota_critical_threshold']

        platform_settings.save()

        # Audit log
        AuditLogger.log_from_request(
            action=AuditAction.PLATFORM_SETTINGS_UPDATE,
            request=request,
            success=True,
            resource_type='PlatformSettings',
            details={'updated': 'Google API quota settings'}
        )

        return Response({
            'message': 'Quota settings updated successfully',
            'settings': {
                'google_tasks_daily_quota': platform_settings.google_tasks_daily_quota,
                'google_calendar_daily_quota': platform_settings.google_calendar_daily_quota,
                'google_drive_daily_quota': platform_settings.google_drive_daily_quota,
                'google_gmail_daily_quota': platform_settings.google_gmail_daily_quota,
                'quota_warning_threshold': platform_settings.quota_warning_threshold,
                'quota_critical_threshold': platform_settings.quota_critical_threshold,
            }
        })


# =============================================================================
# REPORT CONFIGURATION VIEWS
# =============================================================================

class ReportConfigurationViewSet(TenantViewSetMixin, viewsets.ModelViewSet):
    """ViewSet for managing report configurations"""
    queryset = ReportConfiguration.objects.all()
    serializer_class = ReportConfigurationSerializer
    permission_classes = [permissions.IsAuthenticated, IsAdminOrReadOnly]

    @action(detail=True, methods=['post'])
    def send_now(self, request, pk=None):
        """Manually trigger a report to be sent immediately"""
        from .services.report_service import ReportService

        report_config = self.get_object()

        if not report_config.is_active:
            return Response(
                {'error': 'Report configuration is not active'},
                status=status.HTTP_400_BAD_REQUEST
            )

        recipients = report_config.get_recipient_list()
        if not recipients:
            return Response(
                {'error': 'No recipients configured'},
                status=status.HTTP_400_BAD_REQUEST
            )

        success, error = ReportService.generate_and_send_report(report_config)

        if success:
            return Response({
                'message': 'Report generated and sent successfully',
                'recipients': recipients,
                'last_sent_at': report_config.last_sent_at
            })
        else:
            return Response(
                {'error': f'Failed to send report: {error}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    @action(detail=True, methods=['get'])
    def preview(self, request, pk=None):
        """Preview report data without sending"""
        from .services.report_service import ReportService

        report_config = self.get_object()
        start_date, end_date = report_config.get_period_dates()

        data = ReportService.get_report_data(
            report_config.organization,
            start_date,
            end_date
        )

        # Convert dates to strings for JSON response
        summary = data['summary']
        summary['period_start'] = summary['period_start'].isoformat()
        summary['period_end'] = summary['period_end'].isoformat()

        return Response({
            'report_config': ReportConfigurationSerializer(report_config).data,
            'data': data
        })

    @action(detail=True, methods=['get'])
    def download_pdf(self, request, pk=None):
        """Download the report as PDF without sending email"""
        from .services.report_service import ReportService
        from django.http import FileResponse

        report_config = self.get_object()
        start_date, end_date = report_config.get_period_dates()

        data = ReportService.get_report_data(
            report_config.organization,
            start_date,
            end_date
        )

        pdf_buffer = ReportService.generate_pdf_report(report_config, data)

        org_name = report_config.organization.name if report_config.organization else 'Report'
        filename = f"{org_name.replace(' ', '_')}_Report_{timezone.now().strftime('%Y%m%d')}.pdf"

        response = FileResponse(
            pdf_buffer,
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['post'])
    def generate_adhoc_pdf(self, request):
        """Generate and download an ad-hoc report PDF"""
        from .services.report_service import ReportService
        from django.http import FileResponse
        from datetime import datetime

        # Get parameters from request
        report_type = request.data.get('report_type', 'TASK_SUMMARY')
        start_date_str = request.data.get('start_date')
        end_date_str = request.data.get('end_date')
        filters = {
            'client_id': request.data.get('client', 'ALL'),
            'work_type_id': request.data.get('work_type', 'ALL'),
            'status': request.data.get('status', 'ALL'),
            'assigned_to': request.data.get('assigned_to', 'ALL'),
        }

        # Parse dates
        start_date = None
        end_date = None
        if start_date_str:
            try:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        if end_date_str:
            try:
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass

        # Get organization from user
        organization = request.user.organization

        # Fetch report data
        data = ReportService.get_adhoc_report_data(
            organization,
            start_date,
            end_date,
            report_type,
            filters
        )

        # Generate PDF
        pdf_buffer = ReportService.generate_adhoc_pdf_report(
            organization,
            report_type,
            data
        )

        # Report type labels for filename
        report_type_labels = {
            'TASK_SUMMARY': 'Task_Summary',
            'CLIENT_SUMMARY': 'Client_Summary',
            'WORK_TYPE_SUMMARY': 'Work_Type_Summary',
            'STAFF_PRODUCTIVITY': 'Staff_Productivity',
            'STATUS_ANALYSIS': 'Status_Analysis',
        }

        org_name = organization.name if organization else 'Report'
        report_label = report_type_labels.get(report_type, 'Adhoc')
        filename = f"{org_name.replace(' ', '_')}_{report_label}_{timezone.now().strftime('%Y%m%d_%H%M')}.pdf"

        response = FileResponse(
            pdf_buffer,
            content_type='application/pdf'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# =============================================================================
# GOOGLE SYNC HUB VIEWSET
# =============================================================================

class GoogleSyncHubViewSet(viewsets.ViewSet):
    """
    ViewSet for Google Sync Hub functionality.
    Handles OAuth flow, sync operations, and settings management.
    """
    permission_classes = [permissions.IsAuthenticated]

    @action(detail=False, methods=['get'])
    def connection_status(self, request):
        """Get current user's Google connection status and settings."""
        user = request.user
        organization = user.organization

        connection, created = GoogleConnection.objects.get_or_create(
            user=user, defaults={'organization': organization}
        )

        sync_settings = None
        if organization:
            sync_settings, _ = GoogleSyncSettings.objects.get_or_create(organization=organization)

        return Response({
            'connection': GoogleConnectionSerializer(connection).data,
            'sync_settings': GoogleSyncSettingsSerializer(sync_settings).data if sync_settings else None,
            'is_admin': user.role in ['ADMIN', 'PARTNER'],
        })

    def _get_google_redirect_uri(self, request):
        """
        Get the redirect URI for Google OAuth.
        Uses frontend URL for better UX (auto-capture code from URL).
        """
        # Get frontend URL from settings or use default
        from django.conf import settings
        frontend_url = getattr(settings, 'FRONTEND_URL', 'http://localhost:3000')
        return f"{frontend_url}/dashboard/google-sync"

    @action(detail=False, methods=['get'])
    def auth_url(self, request):
        """Generate Google OAuth authorization URL."""
        from .services.google_oauth_service import GoogleOAuthService
        try:
            redirect_uri = self._get_google_redirect_uri(request)
            auth_url, state = GoogleOAuthService.get_authorization_url(redirect_uri=redirect_uri, state=str(request.user.id))
            return Response({'auth_url': auth_url, 'state': state})
        except ValueError as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def connect(self, request):
        """Complete OAuth flow with authorization code."""
        from .services.google_oauth_service import GoogleOAuthService
        code = request.data.get('code')
        if not code:
            return Response({'error': 'Authorization code is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            redirect_uri = self._get_google_redirect_uri(request)
            credentials = GoogleOAuthService.exchange_code_for_tokens(code, redirect_uri)
            connection, _ = GoogleConnection.objects.get_or_create(user=request.user, defaults={'organization': request.user.organization})
            GoogleOAuthService.save_credentials_to_connection(connection, credentials)
            return Response({'success': True, 'connection': GoogleConnectionSerializer(connection).data})
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def disconnect(self, request):
        """Disconnect Google account."""
        from .services.google_oauth_service import GoogleOAuthService
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            GoogleOAuthService.disconnect(connection)
            return Response({'success': True, 'message': 'Google account disconnected successfully'})
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def update_services(self, request):
        """Update enabled Google services (tasks, calendar, drive, gmail)."""
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            if connection.status != 'CONNECTED':
                return Response({'error': 'Google account not connected'}, status=status.HTTP_400_BAD_REQUEST)
            serializer = GoogleConnectionUpdateSerializer(connection, data=request.data, partial=True)
            if serializer.is_valid():
                serializer.save()
                return Response({'success': True, 'connection': GoogleConnectionSerializer(connection).data})
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get', 'put'], url_path='sync-settings')
    def sync_settings(self, request):
        """Get or update organization Google sync settings (Admin only)."""
        user = request.user
        if user.role not in ['ADMIN', 'PARTNER']:
            return Response({'error': 'Only admins can manage sync settings'}, status=status.HTTP_403_FORBIDDEN)
        sync_settings, _ = GoogleSyncSettings.objects.get_or_create(organization=user.organization)
        if request.method == 'GET':
            return Response(GoogleSyncSettingsSerializer(sync_settings).data)
        serializer = GoogleSyncSettingsSerializer(sync_settings, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    @action(detail=False, methods=['post'])
    def sync_all_tasks(self, request):
        """Manually sync all tasks to Google Tasks and/or Calendar."""
        from .services.google_tasks_service import GoogleTasksService
        from .services.google_calendar_service import GoogleCalendarService
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            if connection.status != 'CONNECTED':
                return Response({'error': 'Google account not connected'}, status=status.HTTP_400_BAD_REQUEST)
            results = {'tasks': {'synced': 0, 'errors': 0}, 'calendar': {'synced': 0, 'errors': 0}}
            if connection.tasks_enabled:
                tasks_service = GoogleTasksService(connection)
                synced, errors = tasks_service.sync_all_tasks()
                results['tasks'] = {'synced': synced, 'errors': errors}
            if connection.calendar_enabled:
                calendar_service = GoogleCalendarService(connection)
                synced, errors = calendar_service.sync_all_tasks()
                results['calendar'] = {'synced': synced, 'errors': errors}
            return Response({'success': True, 'results': results})
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['post'])
    def sync_task(self, request):
        """Sync a specific task to Google Tasks and/or Calendar."""
        from .services.google_tasks_service import GoogleTasksService
        from .services.google_calendar_service import GoogleCalendarService
        task_id = request.data.get('task_id')
        if not task_id:
            return Response({'error': 'task_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            work_instance = WorkInstance.objects.get(id=task_id, organization=request.user.organization)
            if connection.status != 'CONNECTED':
                return Response({'error': 'Google account not connected'}, status=status.HTTP_400_BAD_REQUEST)
            results = {}
            if connection.tasks_enabled:
                tasks_service = GoogleTasksService(connection)
                mapping = tasks_service.sync_task_to_google(work_instance)
                results['task_mapping'] = GoogleTaskMappingSerializer(mapping).data
            if connection.calendar_enabled:
                calendar_service = GoogleCalendarService(connection)
                mapping = calendar_service.sync_task_to_calendar(work_instance)
                results['calendar_mapping'] = GoogleCalendarMappingSerializer(mapping).data
            return Response({'success': True, 'results': results})
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)
        except WorkInstance.DoesNotExist:
            return Response({'error': 'Task not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def sync_logs(self, request):
        """Get sync logs for the current user."""
        logs = GoogleSyncLog.objects.filter(user=request.user).order_by('-created_at')[:100]
        return Response(GoogleSyncLogSerializer(logs, many=True).data)

    @action(detail=False, methods=['get'])
    def task_mappings(self, request):
        """Get all task mappings for the current user."""
        mappings = GoogleTaskMapping.objects.filter(user=request.user).select_related('work_instance__client_work__client', 'work_instance__client_work__work_type')
        return Response(GoogleTaskMappingSerializer(mappings, many=True).data)

    @action(detail=False, methods=['get'])
    def calendar_mappings(self, request):
        """Get all calendar mappings for the current user."""
        mappings = GoogleCalendarMapping.objects.filter(user=request.user).select_related('work_instance__client_work__client', 'work_instance__client_work__work_type')
        return Response(GoogleCalendarMappingSerializer(mappings, many=True).data)

    @action(detail=False, methods=['get'])
    def drive_folders(self, request):
        """Get all Drive folder mappings for the organization."""
        mappings = GoogleDriveMapping.objects.filter(organization=request.user.organization).select_related('client')
        return Response(GoogleDriveMappingSerializer(mappings, many=True).data)

    @action(detail=False, methods=['post'])
    def create_client_folders(self, request):
        """Create Google Drive folder structure for a client."""
        from .services.google_drive_service import GoogleDriveService
        client_id = request.data.get('client_id')
        if not client_id:
            return Response({'error': 'client_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            client = Client.objects.get(id=client_id, organization=request.user.organization)
            if connection.status != 'CONNECTED' or not connection.drive_enabled:
                return Response({'error': 'Google Drive not enabled'}, status=status.HTTP_400_BAD_REQUEST)
            drive_service = GoogleDriveService(connection)
            mapping = drive_service.create_client_folder_structure(client)
            return Response({'success': True, 'mapping': GoogleDriveMappingSerializer(mapping).data})
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)
        except Client.DoesNotExist:
            return Response({'error': 'Client not found'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'])
    def task_lists(self, request):
        """Get available Google Task lists."""
        from .services.google_oauth_service import GoogleOAuthService
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            if connection.status != 'CONNECTED':
                return Response({'error': 'Google account not connected'}, status=status.HTTP_400_BAD_REQUEST)
            credentials = GoogleOAuthService.get_credentials_from_connection(connection)
            task_lists = GoogleOAuthService.get_task_lists(credentials)
            return Response({'task_lists': task_lists, 'current_list_id': connection.tasks_list_id})
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['get'])
    def calendars(self, request):
        """Get available Google Calendars."""
        from .services.google_oauth_service import GoogleOAuthService
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            if connection.status != 'CONNECTED':
                return Response({'error': 'Google account not connected'}, status=status.HTTP_400_BAD_REQUEST)
            credentials = GoogleOAuthService.get_credentials_from_connection(connection)
            calendars = GoogleOAuthService.get_calendars(credentials)
            return Response({'calendars': calendars, 'current_calendar_id': connection.calendar_id})
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def set_task_list(self, request):
        """Set the default Google Task list for syncing."""
        task_list_id = request.data.get('task_list_id')
        if not task_list_id:
            return Response({'error': 'task_list_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            connection.tasks_list_id = task_list_id
            connection.save(update_fields=['tasks_list_id'])
            return Response({'success': True, 'connection': GoogleConnectionSerializer(connection).data})
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=False, methods=['post'])
    def set_calendar(self, request):
        """Set the default Google Calendar for syncing."""
        calendar_id = request.data.get('calendar_id')
        if not calendar_id:
            return Response({'error': 'calendar_id is required'}, status=status.HTTP_400_BAD_REQUEST)
        try:
            connection = GoogleConnection.objects.get(user=request.user)
            connection.calendar_id = calendar_id
            connection.save(update_fields=['calendar_id'])
            return Response({'success': True, 'connection': GoogleConnectionSerializer(connection).data})
        except GoogleConnection.DoesNotExist:
            return Response({'error': 'No Google connection found'}, status=status.HTTP_404_NOT_FOUND)
